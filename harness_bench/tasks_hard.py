"""Tasks 101..150 — harder benchmark wave.

These tasks intentionally force the agent to do more than text edits:

- compute non-trivial aggregates (mean / sum / group-by) on dozens of rows;
- work with Excel (.xlsx) and SQLite (.db) — binary file formats;
- implement Python functions and either run them or pass pytest tests;
- locate patterns across 10+ files where reading each by hand is tedious
  (so `grep` and `glob` are the natural tools);
- parse access logs, multi-file projects, JSONL streams, YAML/INI configs.

Binary files (xlsx, sqlite) are produced via `setup_callback` / `gold_callback`
hooks on `Task` (see `core.py`). Text fixtures are generated programmatically
at module load so the setup, gold, and verifier all share the same data.
"""

from __future__ import annotations

import csv
import hashlib
import io
import json
import re
import sqlite3
from collections import Counter
from pathlib import Path

import yaml

from harness_bench.core import Task, VerifyResult
from harness_bench.verifiers import (
    all_of,
    file_contains,
    file_exists,
    file_lines_equal,
    file_matches_regex,
    file_text_equals,
    pytest_passes,
    python_callable_returns,
    xlsx_cell_equals,
)

# ---------------------------------------------------------------------------
# Group A: CSV (10 tasks, 101..110)
# ---------------------------------------------------------------------------

# 101. csv_mean_score — average over 30 rows
_SCORES_LIST = list(range(50, 80))  # mean = 64.5
_SCORES_CSV = "name,score\n" + "".join(
    f"player_{i + 1:02d},{s}\n" for i, s in enumerate(_SCORES_LIST)
)
TASK_101 = Task(
    id="task_101_csv_mean_score",
    name="Compute mean of the score column",
    tags=("csv", "compute", "execute", "medium"),
    prompt=(
        "В файле scores.csv 30 строк данных (после строки-заголовка name,score)."
        " Посчитай среднее арифметическое по столбцу score и запиши его одной"
        " строкой в файл mean.txt, округлив до двух знаков после запятой (точка"
        " как десятичный разделитель, например '64.50')."
    ),
    setup_files={"scores.csv": _SCORES_CSV},
    gold_files={"mean.txt": "64.50\n"},
    verifier=file_text_equals("mean.txt", "64.50"),
)


# 102. csv_filter_adults — keep rows with age >= 18
_AGES = [
    15, 22, 17, 25, 16, 19, 31, 14, 18, 27, 33, 16, 21, 17, 28, 19, 15, 30, 17, 22,
    18, 16, 24, 17, 20, 19, 16, 23, 18, 17, 25, 15, 19, 17, 22, 18, 16, 21, 17, 19,
]
_USERS_CSV = "name,age\n" + "".join(f"user_{i + 1:02d},{a}\n" for i, a in enumerate(_AGES))
_ADULTS_CSV = "name,age\n" + "".join(
    f"user_{i + 1:02d},{a}\n" for i, a in enumerate(_AGES) if a >= 18
)
TASK_102 = Task(
    id="task_102_csv_filter_adults",
    name="Keep only adults (age>=18) from users.csv",
    tags=("csv", "filter", "execute", "medium"),
    prompt=(
        "В файле users.csv 40 строк данных в формате name,age (плюс строка-"
        "заголовок). Сформируй файл adults.csv: в нём должна быть та же строка-"
        "заголовок, а затем только строки, у которых age >= 18 — в исходном"
        " порядке. Не меняй формат CSV."
    ),
    setup_files={"users.csv": _USERS_CSV},
    gold_files={"adults.csv": _ADULTS_CSV},
    verifier=file_text_equals("adults.csv", _ADULTS_CSV),
)


# 103. csv_sort_desc — sort by score column descending
_SORT_INPUT_ROWS = [
    ("alpha", 42), ("bravo", 88), ("charlie", 15), ("delta", 73), ("echo", 91),
    ("foxtrot", 27), ("golf", 64), ("hotel", 50), ("india", 36), ("juliet", 80),
    ("kilo", 58), ("lima", 19), ("mike", 95), ("november", 71), ("oscar", 33),
]
_SORT_INPUT_CSV = "team,score\n" + "".join(f"{n},{s}\n" for n, s in _SORT_INPUT_ROWS)
_SORT_GOLD_ROWS = sorted(_SORT_INPUT_ROWS, key=lambda r: -r[1])
_SORT_GOLD_CSV = "team,score\n" + "".join(f"{n},{s}\n" for n, s in _SORT_GOLD_ROWS)
TASK_103 = Task(
    id="task_103_csv_sort_desc",
    name="Sort teams.csv by score descending",
    tags=("csv", "sort", "execute", "medium"),
    prompt=(
        "В файле teams.csv 15 строк данных в формате team,score (плюс строка-"
        "заголовок). Отсортируй строки данных по столбцу score по убыванию и"
        " сохрани результат в sorted.csv (включая ту же строку-заголовок)."
        " Названия команд и значения не меняй."
    ),
    setup_files={"teams.csv": _SORT_INPUT_CSV},
    gold_files={"sorted.csv": _SORT_GOLD_CSV},
    verifier=file_text_equals("sorted.csv", _SORT_GOLD_CSV),
)


# 104. csv_join — inner join by user_id
_JOIN_USERS = "user_id,name\n1,Alice\n2,Bob\n3,Carol\n4,Dave\n5,Eve\n"
_JOIN_ORDERS = "user_id,amount\n1,100\n3,250\n3,40\n5,90\n2,75\n5,10\n"
# Inner join preserving order from orders.csv:
_JOIN_GOLD = (
    "user_id,name,amount\n"
    "1,Alice,100\n"
    "3,Carol,250\n"
    "3,Carol,40\n"
    "5,Eve,90\n"
    "2,Bob,75\n"
    "5,Eve,10\n"
)
TASK_104 = Task(
    id="task_104_csv_join",
    name="Inner-join users.csv and orders.csv by user_id",
    tags=("csv", "join", "execute", "hard"),
    prompt=(
        "В файлах users.csv (заголовок user_id,name) и orders.csv (заголовок"
        " user_id,amount) есть данные. Сделай inner join по user_id и сохрани"
        " результат в joined.csv с заголовком 'user_id,name,amount'. Сохрани"
        " порядок строк такой же, как в orders.csv (для каждой строки из"
        " orders.csv добавь поле name из соответствующей записи в users.csv)."
    ),
    setup_files={"users.csv": _JOIN_USERS, "orders.csv": _JOIN_ORDERS},
    gold_files={"joined.csv": _JOIN_GOLD},
    verifier=file_text_equals("joined.csv", _JOIN_GOLD),
)


# 105. csv_to_tsv — convert csv to tsv
_C2T_CSV = "name,age,city\nAlice,30,Moscow\nBob,25,Berlin\nCarol,40,Paris\n"
_C2T_TSV = "name\tage\tcity\nAlice\t30\tMoscow\nBob\t25\tBerlin\nCarol\t40\tParis\n"
TASK_105 = Task(
    id="task_105_csv_to_tsv",
    name="Convert data.csv to TSV (tab-separated)",
    tags=("csv", "convert", "execute", "easy"),
    prompt=(
        "Преобразуй файл data.csv (запятая как разделитель) в data.tsv (символ"
        " табуляции как разделитель). Заголовки и значения сохрани, порядок"
        " строк не меняй."
    ),
    setup_files={"data.csv": _C2T_CSV},
    gold_files={"data.tsv": _C2T_TSV},
    verifier=file_text_equals("data.tsv", _C2T_TSV),
)


# 106. csv_group_count — group sales by category
_SALES_CATS = ["food", "tech", "tech", "books", "food", "tech", "books", "food",
               "tech", "food", "tech", "tech", "books", "food", "tech",
               "books", "food", "tech", "tech", "books"]
_SALES_CSV = "id,category\n" + "".join(f"{i + 1},{c}\n" for i, c in enumerate(_SALES_CATS))
_SALES_COUNTS = Counter(_SALES_CATS)
_SALES_GOLD = "category,count\n" + "".join(
    f"{k},{_SALES_COUNTS[k]}\n" for k in ["books", "food", "tech"]
)
TASK_106 = Task(
    id="task_106_csv_group_count",
    name="Count sales per category",
    tags=("csv", "groupby", "execute", "medium"),
    prompt=(
        "В файле sales.csv 20 строк данных (id,category). Сформируй файл"
        " counts.csv с заголовком 'category,count' и тремя строками: по одной"
        " на каждую уникальную категорию, со значением — числом её появлений в"
        " sales.csv. Категории расположи в алфавитном порядке по возрастанию."
    ),
    setup_files={"sales.csv": _SALES_CSV},
    gold_files={"counts.csv": _SALES_GOLD},
    verifier=file_text_equals("counts.csv", _SALES_GOLD),
)


# 107. csv_max_row — row with max amount
_PAY_ROWS = [
    ("INV-001", 120), ("INV-002", 540), ("INV-003", 80), ("INV-004", 1100),
    ("INV-005", 760), ("INV-006", 1100), ("INV-007", 1099), ("INV-008", 970),
    ("INV-009", 230), ("INV-010", 1080),
]
_PAY_CSV = "id,amount\n" + "".join(f"{i},{a}\n" for i, a in _PAY_ROWS)
# First row with max=1100 is INV-004
_PAY_GOLD = "INV-004,1100\n"
TASK_107 = Task(
    id="task_107_csv_max_row",
    name="Pick the row with the largest amount",
    tags=("csv", "compute", "execute", "medium"),
    prompt=(
        "В файле payments.csv 10 строк данных (id,amount). Найди строку с"
        " максимальным значением amount (если максимум встречается у нескольких"
        " строк — возьми первую такую по порядку из payments.csv) и запиши её"
        " как есть (без заголовка) одной строкой в файл max.csv в формате"
        " 'id,amount'."
    ),
    setup_files={"payments.csv": _PAY_CSV},
    gold_files={"max.csv": _PAY_GOLD},
    verifier=file_text_equals("max.csv", "INV-004,1100"),
)


# 108. csv_dedupe — keep first occurrence per user_id
_DEDUPE_INPUT = (
    "user_id,event\n"
    "u1,login\n"
    "u2,login\n"
    "u1,view\n"
    "u3,login\n"
    "u2,click\n"
    "u4,login\n"
    "u1,logout\n"
    "u3,view\n"
    "u5,login\n"
)
_DEDUPE_GOLD = (
    "user_id,event\n"
    "u1,login\n"
    "u2,login\n"
    "u3,login\n"
    "u4,login\n"
    "u5,login\n"
)
TASK_108 = Task(
    id="task_108_csv_dedupe",
    name="Deduplicate events.csv by user_id (first occurrence)",
    tags=("csv", "dedupe", "execute", "medium"),
    prompt=(
        "В файле events.csv колонки user_id,event и несколько событий на одного"
        " пользователя. Сделай файл unique.csv с теми же колонками, в котором"
        " от каждого user_id осталось только первое по порядку событие из"
        " events.csv. Заголовок сохрани."
    ),
    setup_files={"events.csv": _DEDUPE_INPUT},
    gold_files={"unique.csv": _DEDUPE_GOLD},
    verifier=file_text_equals("unique.csv", _DEDUPE_GOLD),
)


# 109. csv_compute_total — add total column = qty * price
_INV_ROWS = [("A", 3, 100), ("B", 5, 20), ("C", 7, 35), ("D", 2, 250), ("E", 10, 9)]
_INV_INPUT = "sku,qty,price\n" + "".join(f"{s},{q},{p}\n" for s, q, p in _INV_ROWS)
_INV_GOLD = "sku,qty,price,total\n" + "".join(
    f"{s},{q},{p},{q * p}\n" for s, q, p in _INV_ROWS
)
TASK_109 = Task(
    id="task_109_csv_add_total",
    name="Add total = qty*price column to invoices.csv",
    tags=("csv", "compute", "execute", "medium"),
    prompt=(
        "В файле invoices.csv колонки sku,qty,price и 5 строк данных. Допиши"
        " столбец total, равный произведению qty * price для каждой строки,"
        " и сохрани результат в том же файле invoices.csv. Новый заголовок —"
        " 'sku,qty,price,total'; порядок строк сохрани."
    ),
    setup_files={"invoices.csv": _INV_INPUT},
    gold_files={"invoices.csv": _INV_GOLD},
    verifier=file_text_equals("invoices.csv", _INV_GOLD),
)


# 110. csv_count_above — count rows where temperature > threshold
_TEMPS = [12, 18, 25, 31, 8, 22, 27, 19, 33, 16, 21, 24, 30, 14, 26, 29, 11, 20, 28, 23]
_TEMPS_CSV = "day,temp\n" + "".join(f"day_{i + 1:02d},{t}\n" for i, t in enumerate(_TEMPS))
_TEMPS_THRESHOLD = 22
_TEMPS_GOLD_COUNT = sum(1 for t in _TEMPS if t > _TEMPS_THRESHOLD)
TASK_110 = Task(
    id="task_110_csv_count_above",
    name="Count rows where temp > 22",
    tags=("csv", "compute", "execute", "easy"),
    prompt=(
        "В файле temps.csv 20 строк данных (day,temp). Посчитай количество"
        " строк, у которых temp > 22, и запиши получившееся число одной строкой"
        " в файл count.txt."
    ),
    setup_files={"temps.csv": _TEMPS_CSV},
    gold_files={"count.txt": f"{_TEMPS_GOLD_COUNT}\n"},
    verifier=file_text_equals("count.txt", str(_TEMPS_GOLD_COUNT)),
)


# ---------------------------------------------------------------------------
# Group B: Excel (3 tasks, 111..113)
# ---------------------------------------------------------------------------


def _xlsx_111_setup(ws: Path) -> None:
    import openpyxl  # noqa: PLC0415

    wb = openpyxl.Workbook()
    sh = wb.active
    sh.title = "Report"
    sh["A1"] = "metric"
    sh["B1"] = "value"
    sh["A2"] = "revenue"
    sh["B2"] = 12345
    sh["A3"] = "cost"
    sh["B3"] = 6700
    wb.save(ws / "report.xlsx")


TASK_111 = Task(
    id="task_111_xlsx_extract_b2",
    name="Extract cell B2 from report.xlsx to value.txt",
    tags=("xlsx", "extract", "execute", "medium"),
    prompt=(
        "В корне рабочей директории лежит файл report.xlsx. Прочитай значение"
        " ячейки B2 на листе Report и запиши его одной строкой в файл value.txt"
        " (как число без кавычек, без лишнего текста). Для чтения xlsx можно"
        " использовать библиотеку openpyxl: она установлена в текущем окружении."
    ),
    setup_files={},
    setup_callback=_xlsx_111_setup,
    gold_files={"value.txt": "12345\n"},
    verifier=file_text_equals("value.txt", "12345"),
)


def _xlsx_112_setup(ws: Path) -> None:
    import openpyxl  # noqa: PLC0415

    wb = openpyxl.Workbook()
    sh = wb.active
    sh.title = "Finances"
    sh.append(["item", "amount"])
    amounts = [120, 80, 250, 75, 1000, 45, 320, 90, 660, 15]
    for i, amt in enumerate(amounts, start=1):
        sh.append([f"item_{i}", amt])
    wb.save(ws / "finances.xlsx")


_XLSX_112_TOTAL = sum([120, 80, 250, 75, 1000, 45, 320, 90, 660, 15])  # 2655
TASK_112 = Task(
    id="task_112_xlsx_sum_column",
    name="Sum the `amount` column from finances.xlsx",
    tags=("xlsx", "compute", "execute", "medium"),
    prompt=(
        "В файле finances.xlsx (лист Finances) есть колонки item и amount, 10"
        " строк данных. Посчитай сумму всех значений в колонке amount и запиши"
        " получившееся число одной строкой в файл total.txt. Используй"
        " openpyxl — он установлен."
    ),
    setup_files={},
    setup_callback=_xlsx_112_setup,
    gold_files={"total.txt": f"{_XLSX_112_TOTAL}\n"},
    verifier=file_text_equals("total.txt", str(_XLSX_112_TOTAL)),
)


def _xlsx_113_setup(ws: Path) -> None:
    import openpyxl  # noqa: PLC0415

    wb = openpyxl.Workbook()
    sh = wb.active
    sh.title = "Inventory"
    sh.append(["sku", "qty", "price"])
    sh.append(["A1", 10, 5])
    sh.append(["B2", 4, 7])
    sh.append(["C3", 8, 9])
    wb.save(ws / "inventory.xlsx")


def _xlsx_113_gold_callback(ws: Path) -> None:
    import openpyxl  # noqa: PLC0415

    wb = openpyxl.load_workbook(ws / "inventory.xlsx")
    sh = wb["Inventory"]
    sh["C3"] = 999
    wb.save(ws / "inventory.xlsx")


TASK_113 = Task(
    id="task_113_xlsx_update_cell",
    name="Set cell C3 in inventory.xlsx to 999",
    tags=("xlsx", "edit", "execute", "medium"),
    prompt=(
        "В файле inventory.xlsx (лист Inventory) есть таблица с колонками"
        " sku, qty, price. Замени значение ячейки C3 на число 999 (остальные"
        " ячейки оставь без изменений) и сохрани файл по тому же пути."
        " Используй openpyxl — он установлен."
    ),
    setup_files={},
    setup_callback=_xlsx_113_setup,
    gold_callback=_xlsx_113_gold_callback,
    verifier=xlsx_cell_equals("inventory.xlsx", "Inventory", "C3", 999),
)


# ---------------------------------------------------------------------------
# Group C: JSON / JSONL (5 tasks, 114..118)
# ---------------------------------------------------------------------------

_TX_AMOUNTS = [10, 20, 30, 40, 50, 60, 70, 80, 90, 100, 11, 22, 33, 44, 55]  # sum = 715
_TX_JSONL = "".join(
    json.dumps({"id": i + 1, "amount": a}) + "\n" for i, a in enumerate(_TX_AMOUNTS)
)
_TX_TOTAL = sum(_TX_AMOUNTS)  # 715
TASK_114 = Task(
    id="task_114_jsonl_sum_amount",
    name="Sum `amount` across JSONL records",
    tags=("jsonl", "compute", "execute", "medium"),
    prompt=(
        "В файле transactions.jsonl 15 строк, каждая — отдельный JSON-объект с"
        " полями id и amount. Посчитай сумму всех значений amount и запиши"
        " число одной строкой в файл total.txt."
    ),
    setup_files={"transactions.jsonl": _TX_JSONL},
    gold_files={"total.txt": f"{_TX_TOTAL}\n"},
    verifier=file_text_equals("total.txt", str(_TX_TOTAL)),
)


_NESTED_CONFIG = json.dumps(
    {
        "service": "api",
        "config": {
            "database": {
                "host": "db.local",
                "credentials": {"user": "postgres", "password": "secret"},
            },
            "cache": {"ttl": 60},
        },
    },
    indent=2,
) + "\n"
TASK_115 = Task(
    id="task_115_json_extract_nested",
    name="Extract config.database.credentials.user from JSON",
    tags=("json", "extract", "execute", "medium"),
    prompt=(
        "В файле config.json есть вложенный JSON-объект. Извлеки значение по"
        " пути config.database.credentials.user и запиши его одной строкой в"
        " файл user.txt (только само значение, без кавычек и без других"
        " символов)."
    ),
    setup_files={"config.json": _NESTED_CONFIG},
    gold_files={"user.txt": "postgres\n"},
    verifier=file_text_equals("user.txt", "postgres"),
)


_FILTER_USERS = [
    {"id": 1, "name": "Alice", "active": True},
    {"id": 2, "name": "Bob", "active": False},
    {"id": 3, "name": "Carol", "active": True},
    {"id": 4, "name": "Dave", "active": False},
    {"id": 5, "name": "Eve", "active": True},
    {"id": 6, "name": "Frank", "active": False},
]
_FILTER_USERS_JSON = json.dumps(_FILTER_USERS, indent=2) + "\n"
_FILTER_ACTIVE_JSON = json.dumps(
    [u for u in _FILTER_USERS if u["active"]], indent=2
) + "\n"


def _verify_task_116(ws: Path) -> VerifyResult:
    p = ws / "active.json"
    if not p.exists():
        return VerifyResult(False, "active.json missing")
    try:
        data = json.loads(p.read_text())
    except json.JSONDecodeError as exc:
        return VerifyResult(False, f"active.json invalid JSON: {exc}")
    if not isinstance(data, list):
        return VerifyResult(False, f"active.json must be a list, got {type(data).__name__}")
    expected = [u for u in _FILTER_USERS if u["active"]]
    if data == expected:
        return VerifyResult(True, "active.json has the expected three records")
    return VerifyResult(
        False, f"active.json content {data!r} differs from expected {expected!r}"
    )


TASK_116 = Task(
    id="task_116_json_filter_active",
    name="Filter active users from users.json",
    tags=("json", "filter", "execute", "medium"),
    prompt=(
        "В файле users.json лежит JSON-массив из 6 объектов с полями id, name,"
        " active (boolean). Сохрани в файл active.json только те объекты, у"
        " которых active == true — в исходном порядке. Это должен быть валидный"
        " JSON-массив."
    ),
    setup_files={"users.json": _FILTER_USERS_JSON},
    gold_files={"active.json": _FILTER_ACTIVE_JSON},
    verifier=_verify_task_116,
)


_J2Y_DATA = {"name": "demo", "version": 3, "debug": False}
_J2Y_JSON = json.dumps(_J2Y_DATA, indent=2) + "\n"


def _verify_task_117(ws: Path) -> VerifyResult:
    p = ws / "data.yaml"
    if not p.exists():
        return VerifyResult(False, "data.yaml missing")
    try:
        data = yaml.safe_load(p.read_text())
    except yaml.YAMLError as exc:
        return VerifyResult(False, f"data.yaml invalid YAML: {exc}")
    if data == _J2Y_DATA:
        return VerifyResult(True, "data.yaml decodes to the expected mapping")
    return VerifyResult(
        False, f"data.yaml decodes to {data!r}, expected {_J2Y_DATA!r}"
    )


TASK_117 = Task(
    id="task_117_json_to_yaml",
    name="Convert data.json to data.yaml",
    tags=("json", "yaml", "convert", "execute", "medium"),
    prompt=(
        "Преобразуй содержимое файла data.json (валидный JSON-объект) в формат"
        " YAML и сохрани результат в файл data.yaml. Структура данных должна"
        " остаться той же — при загрузке data.yaml через yaml.safe_load должен"
        " получиться тот же объект, что и при json.loads(data.json)."
    ),
    setup_files={"data.json": _J2Y_JSON},
    gold_files={"data.yaml": yaml.safe_dump(_J2Y_DATA, sort_keys=False)},
    verifier=_verify_task_117,
)


_EVENTS_LIST = [
    {"id": 1, "status": "ok"},
    {"id": 2, "status": "err"},
    {"id": 3, "status": "ok"},
    {"id": 4, "status": "ok"},
    {"id": 5, "status": "err"},
    {"id": 6, "status": "warn"},
    {"id": 7, "status": "ok"},
    {"id": 8, "status": "err"},
    {"id": 9, "status": "ok"},
    {"id": 10, "status": "warn"},
]
_EVENTS_JSONL = "".join(json.dumps(e) + "\n" for e in _EVENTS_LIST)
_EVENTS_TALLY = dict(Counter(e["status"] for e in _EVENTS_LIST))  # ok 5, err 3, warn 2


def _verify_task_118(ws: Path) -> VerifyResult:
    p = ws / "tally.json"
    if not p.exists():
        return VerifyResult(False, "tally.json missing")
    try:
        data = json.loads(p.read_text())
    except json.JSONDecodeError as exc:
        return VerifyResult(False, f"tally.json invalid JSON: {exc}")
    if not isinstance(data, dict):
        return VerifyResult(False, f"tally.json must be an object, got {type(data).__name__}")
    if data == _EVENTS_TALLY:
        return VerifyResult(True, "tally.json matches expected counts")
    return VerifyResult(
        False, f"tally.json {data!r} differs from expected {_EVENTS_TALLY!r}"
    )


TASK_118 = Task(
    id="task_118_jsonl_count_by_status",
    name="Tally events by status",
    tags=("jsonl", "groupby", "execute", "medium"),
    prompt=(
        "В файле events.jsonl 10 JSON-объектов (по одному на строку), у каждого"
        " есть поле status со значениями 'ok', 'err' или 'warn'. Сформируй файл"
        " tally.json — валидный JSON-объект, в котором ключи — это значения"
        " status, а значения — количество событий с таким статусом."
    ),
    setup_files={"events.jsonl": _EVENTS_JSONL},
    gold_files={"tally.json": json.dumps(_EVENTS_TALLY) + "\n"},
    verifier=_verify_task_118,
)


# ---------------------------------------------------------------------------
# Group D: YAML / INI / TOML (4 tasks, 119..122)
# ---------------------------------------------------------------------------


_YAML_INITIAL = "name: demo\nversion: '1.0'\ndebug: false\n"


def _verify_task_119(ws: Path) -> VerifyResult:
    p = ws / "config.yaml"
    if not p.exists():
        return VerifyResult(False, "config.yaml missing")
    try:
        data = yaml.safe_load(p.read_text())
    except yaml.YAMLError as exc:
        return VerifyResult(False, f"config.yaml invalid YAML: {exc}")
    if not isinstance(data, dict):
        return VerifyResult(False, f"config.yaml is not a mapping: {data!r}")
    if data.get("name") != "demo":
        return VerifyResult(False, f"name changed: {data.get('name')!r}")
    if data.get("debug") is not False:
        return VerifyResult(False, f"debug changed: {data.get('debug')!r}")
    # version may be string "2.0" or float 2.0 — accept both
    version = data.get("version")
    if version not in ("2.0", 2.0):
        return VerifyResult(False, f"version is {version!r}, expected '2.0' or 2.0")
    return VerifyResult(True, "config.yaml version bumped, other fields intact")


TASK_119 = Task(
    id="task_119_yaml_bump_version",
    name="Bump version field in config.yaml from 1.0 to 2.0",
    tags=("yaml", "edit", "execute", "medium"),
    prompt=(
        "В файле config.yaml (валидный YAML с полями name, version, debug)"
        " смени значение поля version с '1.0' на '2.0' (как строку или число —"
        " неважно). Поля name и debug сохрани в исходном виде. Файл должен"
        " остаться валидным YAML, парсящимся через yaml.safe_load."
    ),
    setup_files={"config.yaml": _YAML_INITIAL},
    gold_files={"config.yaml": "name: demo\nversion: '2.0'\ndebug: false\n"},
    verifier=_verify_task_119,
)


_INI_INITIAL = (
    "[main]\n"
    "name = demo\n"
    "version = 1.0\n"
    "\n"
    "[paths]\n"
    "data = /var/data\n"
)
_INI_GOLD = _INI_INITIAL + "\n[logging]\nlevel = INFO\nfile = app.log\n"


def _verify_task_120(ws: Path) -> VerifyResult:
    import configparser  # noqa: PLC0415

    p = ws / "settings.ini"
    if not p.exists():
        return VerifyResult(False, "settings.ini missing")
    parser = configparser.ConfigParser()
    try:
        parser.read_string(p.read_text())
    except configparser.Error as exc:
        return VerifyResult(False, f"settings.ini invalid INI: {exc}")
    if "logging" not in parser.sections():
        return VerifyResult(False, "settings.ini has no [logging] section")
    if parser["logging"].get("level") != "INFO":
        return VerifyResult(False, f"logging.level={parser['logging'].get('level')!r}")
    if parser["logging"].get("file") != "app.log":
        return VerifyResult(False, f"logging.file={parser['logging'].get('file')!r}")
    if parser["main"].get("name") != "demo":
        return VerifyResult(False, "main.name was changed")
    if parser["paths"].get("data") != "/var/data":
        return VerifyResult(False, "paths.data was changed")
    return VerifyResult(True, "settings.ini has the new [logging] section")


TASK_120 = Task(
    id="task_120_ini_add_section",
    name="Add [logging] section to settings.ini",
    tags=("ini", "edit", "execute", "medium"),
    prompt=(
        "В файле settings.ini уже есть секции [main] и [paths]. Добавь новую"
        " секцию [logging] с двумя ключами:\n"
        "    level = INFO\n"
        "    file = app.log\n"
        " Существующие секции и их значения не меняй. Файл должен остаться"
        " валидным INI, парсящимся через configparser."
    ),
    setup_files={"settings.ini": _INI_INITIAL},
    gold_files={"settings.ini": _INI_GOLD},
    verifier=_verify_task_120,
)


_TOML_INITIAL = (
    "[project]\n"
    'name = "demo"\n'
    'version = "0.1.0"\n'
    'dependencies = ["requests", "httpx"]\n'
)


def _verify_task_121(ws: Path) -> VerifyResult:
    import tomllib  # noqa: PLC0415

    p = ws / "pyproject.toml"
    if not p.exists():
        return VerifyResult(False, "pyproject.toml missing")
    try:
        data = tomllib.loads(p.read_text())
    except tomllib.TOMLDecodeError as exc:
        return VerifyResult(False, f"pyproject.toml invalid TOML: {exc}")
    deps = data.get("project", {}).get("dependencies")
    if not isinstance(deps, list):
        return VerifyResult(False, f"project.dependencies is not a list: {deps!r}")
    if "pydantic" not in deps:
        return VerifyResult(False, f"'pydantic' not in dependencies: {deps!r}")
    if set(deps) != {"requests", "httpx", "pydantic"}:
        return VerifyResult(
            False, f"dependencies set mismatch: {sorted(deps)!r}"
        )
    if data["project"].get("name") != "demo":
        return VerifyResult(False, "project.name changed")
    return VerifyResult(True, "pyproject.toml dependencies extended")


TASK_121 = Task(
    id="task_121_toml_add_dep",
    name="Add `pydantic` to project.dependencies in pyproject.toml",
    tags=("toml", "edit", "execute", "medium"),
    prompt=(
        "В файле pyproject.toml в секции [project] поле dependencies — это"
        " массив с двумя строками: 'requests' и 'httpx'. Добавь в него третью"
        " запись 'pydantic'. Поля name и version не меняй; итоговый файл должен"
        " остаться валидным TOML (парситься через tomllib)."
    ),
    setup_files={"pyproject.toml": _TOML_INITIAL},
    gold_files={
        "pyproject.toml": (
            "[project]\n"
            'name = "demo"\n'
            'version = "0.1.0"\n'
            'dependencies = ["requests", "httpx", "pydantic"]\n'
        ),
    },
    verifier=_verify_task_121,
)


# 122. yaml_to_json (reverse of 117)
_Y2J_YAML = "name: gigachat\nversion: 3\ndebug: false\nfeatures:\n  - tools\n  - think\n"
_Y2J_DATA = yaml.safe_load(_Y2J_YAML)


def _verify_task_122(ws: Path) -> VerifyResult:
    p = ws / "config.json"
    if not p.exists():
        return VerifyResult(False, "config.json missing")
    try:
        data = json.loads(p.read_text())
    except json.JSONDecodeError as exc:
        return VerifyResult(False, f"config.json invalid JSON: {exc}")
    if data == _Y2J_DATA:
        return VerifyResult(True, "config.json equals the YAML structure")
    return VerifyResult(False, f"config.json {data!r} differs from expected {_Y2J_DATA!r}")


TASK_122 = Task(
    id="task_122_yaml_to_json",
    name="Convert config.yaml to config.json",
    tags=("yaml", "json", "convert", "execute", "medium"),
    prompt=(
        "Преобразуй файл config.yaml в config.json (валидный JSON). Структура"
        " данных должна сохраниться: после json.loads(config.json) должен"
        " получиться тот же объект, что и yaml.safe_load(config.yaml). Формат"
        " отступов в JSON не важен."
    ),
    setup_files={"config.yaml": _Y2J_YAML},
    gold_files={"config.json": json.dumps(_Y2J_DATA) + "\n"},
    verifier=_verify_task_122,
)


# ---------------------------------------------------------------------------
# Group E: SQLite (2 tasks, 123..124)
# ---------------------------------------------------------------------------


def _sqlite_123_setup(ws: Path) -> None:
    conn = sqlite3.connect(ws / "users.db")
    conn.execute("CREATE TABLE users (id INTEGER PRIMARY KEY, name TEXT, age INTEGER)")
    rows = [
        (1, "Alice", 30), (2, "Bob", 25), (3, "Carol", 40), (4, "Dave", 22),
        (5, "Eve", 33), (6, "Frank", 19), (7, "Grace", 28), (8, "Henry", 45),
        (9, "Ivy", 27), (10, "Jack", 31), (11, "Kate", 24), (12, "Liam", 36),
    ]
    conn.executemany("INSERT INTO users VALUES (?, ?, ?)", rows)
    conn.commit()
    conn.close()


_SQLITE_123_COUNT = 12
TASK_123 = Task(
    id="task_123_sqlite_count",
    name="Count rows in users.db users table",
    tags=("sqlite", "compute", "execute", "medium"),
    prompt=(
        "В корне рабочей директории лежит SQLite-база users.db с таблицей"
        " users (колонки id, name, age). Посчитай количество строк в этой"
        " таблице и запиши число одной строкой в файл count.txt. Используй"
        " стандартную библиотеку sqlite3."
    ),
    setup_files={},
    setup_callback=_sqlite_123_setup,
    gold_files={"count.txt": f"{_SQLITE_123_COUNT}\n"},
    verifier=file_text_equals("count.txt", str(_SQLITE_123_COUNT)),
)


def _sqlite_124_setup(ws: Path) -> None:
    conn = sqlite3.connect(ws / "shop.db")
    conn.execute("CREATE TABLE orders (id INTEGER PRIMARY KEY, amount INTEGER)")
    amounts = [120, 350, 80, 1700, 95, 240, 60, 1100, 220, 480]
    conn.executemany(
        "INSERT INTO orders VALUES (?, ?)",
        [(i + 1, a) for i, a in enumerate(amounts)],
    )
    conn.commit()
    conn.close()


_SQLITE_124_TOTAL = 120 + 350 + 80 + 1700 + 95 + 240 + 60 + 1100 + 220 + 480  # 4445
TASK_124 = Task(
    id="task_124_sqlite_sum",
    name="Sum the orders.amount column from shop.db",
    tags=("sqlite", "compute", "execute", "medium"),
    prompt=(
        "В корне рабочей директории лежит SQLite-база shop.db с таблицей"
        " orders (колонки id, amount). Посчитай сумму всех значений в столбце"
        " amount и запиши число одной строкой в файл total.txt. Используй"
        " стандартную библиотеку sqlite3."
    ),
    setup_files={},
    setup_callback=_sqlite_124_setup,
    gold_files={"total.txt": f"{_SQLITE_124_TOTAL}\n"},
    verifier=file_text_equals("total.txt", str(_SQLITE_124_TOTAL)),
)


# ---------------------------------------------------------------------------
# Group F: Python implementation + execution (10 tasks, 125..134)
# ---------------------------------------------------------------------------

TASK_125 = Task(
    id="task_125_impl_fib",
    name="Implement fib(n) (iterative)",
    tags=("python", "impl", "execute", "medium"),
    prompt=(
        "Создай файл fib.py с функцией fib(n: int) -> int, которая возвращает"
        " n-е число Фибоначчи (fib(0) == 0, fib(1) == 1, fib(2) == 1, fib(10)"
        " == 55, fib(20) == 6765). Реализуй итеративно за O(n) или любым"
        " корректным способом — главное, чтобы возвращаемые значения совпадали"
        " с примерами выше."
    ),
    setup_files={},
    gold_files={
        "fib.py": (
            "def fib(n: int) -> int:\n"
            "    a, b = 0, 1\n"
            "    for _ in range(n):\n"
            "        a, b = b, a + b\n"
            "    return a\n"
        ),
    },
    verifier=all_of(
        python_callable_returns("fib.py", "mod.fib(0)", 0),
        python_callable_returns("fib.py", "mod.fib(1)", 1),
        python_callable_returns("fib.py", "mod.fib(10)", 55),
        python_callable_returns("fib.py", "mod.fib(20)", 6765),
    ),
)


TASK_126 = Task(
    id="task_126_impl_count_vowels",
    name="Implement count_vowels(s)",
    tags=("python", "impl", "execute", "easy"),
    prompt=(
        "Создай файл vowels.py с функцией count_vowels(s: str) -> int, которая"
        " считает количество гласных букв в строке. Считай гласными ровно"
        " следующие 10 символов: a, e, i, o, u и их версии в верхнем регистре"
        " A, E, I, O, U. Другие буквы и не-буквы не учитывай."
    ),
    setup_files={},
    gold_files={
        "vowels.py": (
            "def count_vowels(s: str) -> int:\n"
            '    return sum(1 for ch in s if ch in "aeiouAEIOU")\n'
        ),
    },
    verifier=all_of(
        python_callable_returns("vowels.py", "mod.count_vowels('hello')", 2),
        python_callable_returns("vowels.py", "mod.count_vowels('AEIOU')", 5),
        python_callable_returns("vowels.py", "mod.count_vowels('xyz')", 0),
        python_callable_returns("vowels.py", "mod.count_vowels('Quick brown fox')", 4),
    ),
)


TASK_127 = Task(
    id="task_127_impl_is_palindrome",
    name="Implement is_palindrome(s)",
    tags=("python", "impl", "execute", "easy"),
    prompt=(
        "Создай файл palindrome.py с функцией is_palindrome(s: str) -> bool."
        " Она должна возвращать True, если строка читается одинаково слева"
        " направо и справа налево (с учётом регистра и пробелов), и False"
        " иначе. Примеры: is_palindrome('aba') == True, is_palindrome('abc')"
        " == False, is_palindrome('') == True."
    ),
    setup_files={},
    gold_files={
        "palindrome.py": (
            "def is_palindrome(s: str) -> bool:\n"
            "    return s == s[::-1]\n"
        ),
    },
    verifier=all_of(
        python_callable_returns("palindrome.py", "mod.is_palindrome('aba')", True),
        python_callable_returns("palindrome.py", "mod.is_palindrome('abc')", False),
        python_callable_returns("palindrome.py", "mod.is_palindrome('')", True),
        python_callable_returns(
            "palindrome.py", "mod.is_palindrome('aabbaa')", True
        ),
    ),
)


# 128. Fix a bug so pytest passes
TASK_128 = Task(
    id="task_128_fix_bug",
    name="Fix the buggy `add` so pytest passes",
    tags=("python", "fix", "pytest", "execute", "hard"),
    prompt=(
        "В файле math_lib.py определена функция add(a, b), которая сейчас"
        " возвращает a - b (это баг). В каталоге tests есть pytest-тест"
        " test_add.py, который проверяет add(2, 3) == 5 и add(-1, 1) == 0."
        " Исправь реализацию add в math_lib.py так, чтобы тесты проходили."
        " Запускать pytest сам не обязан — нам важно лишь, чтобы тесты"
        " проходили, когда мы их запустим."
    ),
    setup_files={
        "math_lib.py": "def add(a, b):\n    return a - b\n",
        "tests/test_add.py": (
            "from math_lib import add\n"
            "\n"
            "\n"
            "def test_two_plus_three():\n"
            "    assert add(2, 3) == 5\n"
            "\n"
            "\n"
            "def test_negative_one_plus_one():\n"
            "    assert add(-1, 1) == 0\n"
        ),
    },
    gold_files={"math_lib.py": "def add(a, b):\n    return a + b\n"},
    verifier=pytest_passes("tests"),
)


# 129. Implement function so given pytest tests pass
TASK_129 = Task(
    id="task_129_implement_passing",
    name="Implement Stack so pytest passes",
    tags=("python", "impl", "pytest", "execute", "hard"),
    prompt=(
        "В каталоге tests есть pytest-тест test_stack.py, который проверяет"
        " работу класса Stack из модуля stack (файл stack.py). Создай файл"
        " stack.py с классом Stack, который поддерживает методы push(value),"
        " pop() (возвращает значение и удаляет его), peek() (возвращает"
        " последнее значение без удаления), is_empty() и __len__(). Тесты не"
        " меняй — нам важно, чтобы они проходили."
    ),
    setup_files={
        "tests/test_stack.py": (
            "from stack import Stack\n"
            "\n"
            "\n"
            "def test_push_pop():\n"
            "    s = Stack()\n"
            "    s.push(1)\n"
            "    s.push(2)\n"
            "    assert s.pop() == 2\n"
            "    assert s.pop() == 1\n"
            "\n"
            "\n"
            "def test_peek_and_len():\n"
            "    s = Stack()\n"
            "    assert s.is_empty()\n"
            "    s.push('a')\n"
            "    s.push('b')\n"
            "    assert s.peek() == 'b'\n"
            "    assert len(s) == 2\n"
            "    assert not s.is_empty()\n"
        ),
    },
    gold_files={
        "stack.py": (
            "class Stack:\n"
            "    def __init__(self):\n"
            "        self._items = []\n"
            "\n"
            "    def push(self, value):\n"
            "        self._items.append(value)\n"
            "\n"
            "    def pop(self):\n"
            "        return self._items.pop()\n"
            "\n"
            "    def peek(self):\n"
            "        return self._items[-1]\n"
            "\n"
            "    def is_empty(self):\n"
            "        return not self._items\n"
            "\n"
            "    def __len__(self):\n"
            "        return len(self._items)\n"
        ),
    },
    verifier=pytest_passes("tests"),
)


# 130. Implement extract_phones via regex
TASK_130 = Task(
    id="task_130_impl_extract_phones",
    name="Implement extract_phones(text) so pytest passes",
    tags=("python", "impl", "regex", "pytest", "execute", "hard"),
    prompt=(
        "Создай файл phones.py с функцией extract_phones(text: str) -> list[str]."
        " Она должна возвращать список всех 10-значных подпоследовательностей"
        " вида '\\d{3}-\\d{3}-\\d{4}' (три цифры, дефис, три цифры, дефис,"
        " четыре цифры) в порядке появления в тексте. В каталоге tests лежит"
        " test_phones.py, который проверит работу. Тесты не меняй."
    ),
    setup_files={
        "tests/test_phones.py": (
            "from phones import extract_phones\n"
            "\n"
            "\n"
            "def test_extracts_two_phones():\n"
            "    text = 'Call me at 415-555-1234 or 212-555-9999, please.'\n"
            "    assert extract_phones(text) == ['415-555-1234', '212-555-9999']\n"
            "\n"
            "\n"
            "def test_no_phones():\n"
            "    assert extract_phones('no phones here') == []\n"
        ),
    },
    gold_files={
        "phones.py": (
            "import re\n"
            "\n"
            "\n"
            "def extract_phones(text: str) -> list[str]:\n"
            '    return re.findall(r"\\d{3}-\\d{3}-\\d{4}", text)\n'
        ),
    },
    verifier=pytest_passes("tests"),
)


# 131. Script that counts words in input and writes the count
_WC_INPUT = "lorem ipsum dolor sit amet consectetur adipiscing elit sed do " * 10
_WC_INPUT = _WC_INPUT.strip() + "\n"  # 100 words on one line
_WC_GOLD = "100"
TASK_131 = Task(
    id="task_131_script_wordcount",
    name="Count words in input.txt and write the count",
    tags=("python", "execute", "compute", "medium"),
    prompt=(
        "В файле input.txt лежит текст. Посчитай в нём количество слов"
        " (разделители — любые пробельные символы) и запиши получившееся число"
        " одной строкой в файл count.txt. Можно использовать любой подход —"
        " написать и запустить python-скрипт или сосчитать через bash; нам"
        " важен только результат в count.txt."
    ),
    setup_files={"input.txt": _WC_INPUT},
    gold_files={"count.txt": _WC_GOLD + "\n"},
    verifier=file_text_equals("count.txt", _WC_GOLD),
)


# 132. Compute md5 hash of a file
_MD5_INPUT = "hello world\n"
_MD5_GOLD = hashlib.md5(_MD5_INPUT.encode("utf-8")).hexdigest()  # noqa: S324
TASK_132 = Task(
    id="task_132_md5_hash",
    name="Write MD5 hash of input.txt to hash.txt",
    tags=("python", "execute", "compute", "medium"),
    prompt=(
        "В файле input.txt лежит текст 'hello world\\n' (12 байт, включая"
        " завершающий перевод строки). Посчитай MD5-хеш содержимого этого"
        " файла (как байтов, без какой-либо предобработки) и запиши"
        " получившуюся шестнадцатеричную строку (32 символа в нижнем регистре)"
        " одной строкой в файл hash.txt. Используй модуль hashlib из стандартной"
        " библиотеки."
    ),
    setup_files={"input.txt": _MD5_INPUT},
    gold_files={"hash.txt": _MD5_GOLD + "\n"},
    verifier=file_text_equals("hash.txt", _MD5_GOLD),
)


# 133. Implement factorial
TASK_133 = Task(
    id="task_133_impl_factorial",
    name="Implement factorial(n)",
    tags=("python", "impl", "execute", "easy"),
    prompt=(
        "Создай файл factorial.py с функцией factorial(n: int) -> int, которая"
        " возвращает n! (произведение чисел от 1 до n; factorial(0) == 1,"
        " factorial(5) == 120, factorial(10) == 3628800)."
    ),
    setup_files={},
    gold_files={
        "factorial.py": (
            "def factorial(n: int) -> int:\n"
            "    result = 1\n"
            "    for i in range(2, n + 1):\n"
            "        result *= i\n"
            "    return result\n"
        ),
    },
    verifier=all_of(
        python_callable_returns("factorial.py", "mod.factorial(0)", 1),
        python_callable_returns("factorial.py", "mod.factorial(5)", 120),
        python_callable_returns("factorial.py", "mod.factorial(10)", 3628800),
    ),
)


# 134. Run a hello script and capture its output
_HELLO_GOLD_STDOUT = "Hello from script"
TASK_134 = Task(
    id="task_134_run_and_capture",
    name="Run script and save its stdout",
    tags=("python", "execute", "medium"),
    prompt=(
        "Создай файл greet.py с кодом, который при запуске печатает в стандартный"
        " вывод ровно одну строку: 'Hello from script' (без кавычек, с"
        " переводом строки в конце). После создания файла запусти его и"
        " сохрани его вывод (то, что напечатал print) в файл out.txt. В out.txt"
        " должна остаться эта же одна строка."
    ),
    setup_files={},
    gold_files={
        "greet.py": 'print("Hello from script")\n',
        "out.txt": _HELLO_GOLD_STDOUT + "\n",
    },
    verifier=all_of(
        file_exists("greet.py"),
        file_text_equals("out.txt", _HELLO_GOLD_STDOUT),
    ),
)


# ---------------------------------------------------------------------------
# Group G: grep / multi-file search (10 tasks, 135..144)
# ---------------------------------------------------------------------------


def _make_imports_project() -> dict[str, str]:
    """Create 20 .py files under src/; total imports is deterministic."""
    files = {}
    # Each file gets 1..3 imports, deterministic per index.
    for i in range(20):
        n_imports = (i % 3) + 1  # 1, 2, or 3
        imports = [f"import mod_{j}\n" for j in range(n_imports)]
        files[f"src/file_{i:02d}.py"] = "".join(imports) + f"\n# file {i}\n"
    return files


_IMPORTS_FILES = _make_imports_project()
_IMPORTS_TOTAL = sum(content.count("import ") for content in _IMPORTS_FILES.values())
TASK_135 = Task(
    id="task_135_grep_count_imports",
    name="Count `import` lines across src/*.py",
    tags=("grep", "search", "execute", "medium"),
    prompt=(
        "В каталоге src лежит 20 .py-файлов. Посчитай общее число строк,"
        " начинающихся со слова 'import' (то есть строк вида 'import <что-то>')"
        " во всех этих файлах. Запиши число одной строкой в файл count.txt."
    ),
    setup_files=_IMPORTS_FILES,
    gold_files={"count.txt": f"{_IMPORTS_TOTAL}\n"},
    verifier=file_text_equals("count.txt", str(_IMPORTS_TOTAL)),
)


def _make_todo_project() -> tuple[dict[str, str], set[str]]:
    """15 files; 4 contain TODO. Return setup + names of TODO-bearing files."""
    files = {}
    todo_idx = {3, 7, 11, 14}
    has_todo = set()
    for i in range(15):
        ext = ["txt", "md", "py", "txt"][i % 4]
        name = f"file_{i:02d}.{ext}"
        body = f"content for {name}\n"
        if i in todo_idx:
            body += "TODO: revise me\n"
            has_todo.add(name)
        files[f"project/{name}"] = body
    return files, has_todo


_TODO_FILES, _TODO_NAMES = _make_todo_project()


def _verify_task_136(ws: Path) -> VerifyResult:
    p = ws / "files.txt"
    if not p.exists():
        return VerifyResult(False, "files.txt missing")
    raw = [line.strip() for line in p.read_text().splitlines() if line.strip()]
    # Accept either bare filenames or relative paths starting with project/.
    normalised = {line.split("/")[-1] for line in raw}
    if normalised != _TODO_NAMES:
        return VerifyResult(
            False,
            f"files.txt entries {sorted(normalised)} differ from expected {sorted(_TODO_NAMES)}",
        )
    return VerifyResult(True, "files.txt has the four TODO-bearing files")


TASK_136 = Task(
    id="task_136_grep_files_with_todo",
    name="List files under project/ that contain TODO",
    tags=("grep", "search", "execute", "medium"),
    prompt=(
        "В каталоге project 15 файлов. Найди все файлы, в содержимом которых"
        " встречается подстрока 'TODO', и запиши их имена в файл files.txt по"
        " одному имени на строку. Можно писать как только имя файла (без"
        " префикса каталога), так и относительный путь с 'project/' впереди —"
        " допустимы оба варианта. Порядок не важен."
    ),
    setup_files=_TODO_FILES,
    gold_files={"files.txt": "\n".join(sorted(_TODO_NAMES)) + "\n"},
    verifier=_verify_task_136,
)


def _make_defs_project() -> dict[str, str]:
    files = {}
    for i in range(12):
        n_defs = (i % 4) + 1  # 1..4
        body = "".join(
            f"def func_{i}_{j}():\n    return {j}\n\n\n" for j in range(n_defs)
        )
        files[f"pkg/mod_{i:02d}.py"] = body
    return files


_DEFS_FILES = _make_defs_project()
_DEFS_TOTAL = sum(content.count("def ") for content in _DEFS_FILES.values())
TASK_137 = Task(
    id="task_137_grep_count_defs",
    name="Count `def ` across pkg/*.py",
    tags=("grep", "search", "execute", "medium"),
    prompt=(
        "В каталоге pkg лежит 12 .py-файлов. Посчитай общее число вхождений"
        " подстроки 'def ' (defение функций; именно с пробелом после 'def')"
        " во всех этих файлах. Запиши число одной строкой в файл count.txt."
    ),
    setup_files=_DEFS_FILES,
    gold_files={"count.txt": f"{_DEFS_TOTAL}\n"},
    verifier=file_text_equals("count.txt", str(_DEFS_TOTAL)),
)


def _make_yaml_configs() -> tuple[dict[str, str], set[str]]:
    files = {}
    has_key = {"prod.yaml", "staging.yaml", "secret.yaml"}
    names = ["dev.yaml", "prod.yaml", "qa.yaml", "staging.yaml",
             "test.yaml", "demo.yaml", "secret.yaml", "local.yaml"]
    for n in names:
        body = "host: example.com\nport: 5432\n"
        if n in has_key:
            body += "api_key: deadbeef\n"
        files[f"configs/{n}"] = body
    return files, has_key


_YAML_FILES, _YAML_NAMES_WITH_KEY = _make_yaml_configs()


def _verify_task_138(ws: Path) -> VerifyResult:
    p = ws / "with_api_key.txt"
    if not p.exists():
        return VerifyResult(False, "with_api_key.txt missing")
    raw = [line.strip() for line in p.read_text().splitlines() if line.strip()]
    normalised = {line.split("/")[-1] for line in raw}
    if normalised != _YAML_NAMES_WITH_KEY:
        return VerifyResult(
            False,
            f"with_api_key.txt {sorted(normalised)} differs from expected {sorted(_YAML_NAMES_WITH_KEY)}",
        )
    return VerifyResult(True, "with_api_key.txt has the three configs that mention api_key")


TASK_138 = Task(
    id="task_138_grep_yaml_with_key",
    name="List YAML configs that contain `api_key:`",
    tags=("grep", "yaml", "search", "execute", "medium"),
    prompt=(
        "В каталоге configs 8 yaml-файлов. Найди те, в которых встречается"
        " подстрока 'api_key:' (двоеточие включительно), и сохрани их имена в"
        " файл with_api_key.txt — по одному на строку, имя файла без префикса"
        " каталога допустимо так же, как относительный путь с 'configs/'."
        " Порядок не важен."
    ),
    setup_files=_YAML_FILES,
    gold_files={"with_api_key.txt": "\n".join(sorted(_YAML_NAMES_WITH_KEY)) + "\n"},
    verifier=_verify_task_138,
)


def _make_logs_project() -> tuple[dict[str, str], list[str]]:
    """6 log files; collect all ERROR lines in deterministic order."""
    files: dict[str, str] = {}
    errors: list[str] = []
    payloads = {
        "logs/app1.log": ["INFO start", "ERROR db timeout", "WARN retry", "ERROR conn lost", "INFO stop"],
        "logs/app2.log": ["INFO ready", "INFO ready", "INFO done"],
        "logs/app3.log": ["ERROR auth failed", "INFO logout"],
        "logs/app4.log": ["WARN slow", "WARN slow"],
        "logs/app5.log": ["ERROR disk full"],
        "logs/app6.log": ["INFO ping", "ERROR network glitch", "INFO ping"],
    }
    for name in sorted(payloads):
        lines = payloads[name]
        files[name] = "\n".join(lines) + "\n"
        errors.extend(line for line in lines if line.startswith("ERROR"))
    return files, errors


_LOGS_FILES, _LOG_ERRORS = _make_logs_project()
_LOG_ERR_COUNT = len(_LOG_ERRORS)
TASK_139 = Task(
    id="task_139_grep_log_errors",
    name="Count ERROR lines across logs/*.log",
    tags=("grep", "logs", "search", "execute", "medium"),
    prompt=(
        "В каталоге logs 6 файлов вида app*.log. Посчитай общее количество"
        " строк, начинающихся со слова 'ERROR' (то есть точно с этих пяти"
        " символов), во всех этих файлах. Запиши число одной строкой в файл"
        " errors_count.txt."
    ),
    setup_files=_LOGS_FILES,
    gold_files={"errors_count.txt": f"{_LOG_ERR_COUNT}\n"},
    verifier=file_text_equals("errors_count.txt", str(_LOG_ERR_COUNT)),
)


def _make_emails_project() -> tuple[dict[str, str], set[str]]:
    files = {}
    emails_per_file = {
        "docs/intro.md": ["alice@example.com"],
        "docs/team.md": ["bob@team.org", "carol@team.org"],
        "docs/about.md": [],
        "docs/contact.md": ["support@demo.net"],
        "docs/changelog.md": [],
    }
    all_emails: set[str] = set()
    for name, mails in emails_per_file.items():
        body = f"# {name}\n"
        for m in mails:
            body += f"Contact: {m}\n"
            all_emails.add(m)
        files[name] = body
    return files, all_emails


_EMAILS_FILES, _EMAILS_ALL = _make_emails_project()


def _verify_task_140(ws: Path) -> VerifyResult:
    p = ws / "emails.txt"
    if not p.exists():
        return VerifyResult(False, "emails.txt missing")
    raw = [line.strip() for line in p.read_text().splitlines() if line.strip()]
    if set(raw) != _EMAILS_ALL:
        return VerifyResult(
            False,
            f"emails.txt {sorted(set(raw))} differs from expected {sorted(_EMAILS_ALL)}",
        )
    return VerifyResult(True, "emails.txt has all four unique emails")


TASK_140 = Task(
    id="task_140_grep_emails",
    name="Collect email addresses from docs/*.md",
    tags=("grep", "regex", "search", "execute", "medium"),
    prompt=(
        "В каталоге docs 5 markdown-файлов. Найди в их содержимом все"
        " email-адреса вида 'name@domain' и сохрани их (по одному на строку, без"
        " дубликатов) в файл emails.txt. Порядок не важен; собирай только сами"
        " адреса, без окружающего текста."
    ),
    setup_files=_EMAILS_FILES,
    gold_files={"emails.txt": "\n".join(sorted(_EMAILS_ALL)) + "\n"},
    verifier=_verify_task_140,
)


def _make_class_project() -> tuple[dict[str, str], list[str]]:
    files = {}
    classes_per_file = {
        "lib/a.py": ["Apple", "Apricot"],
        "lib/b.py": ["Banana"],
        "lib/c.py": ["Cherry", "Coconut"],
        "lib/d.py": [],
        "lib/e.py": ["Eggplant"],
        "lib/f.py": ["Fig", "Feijoa"],
    }
    all_classes: list[str] = []
    for name, classes in classes_per_file.items():
        body = "".join(f"class {c}:\n    pass\n\n\n" for c in classes)
        files[name] = body or "# empty\n"
        all_classes.extend(classes)
    return files, sorted(all_classes)


_CLASS_FILES, _CLASS_SORTED = _make_class_project()
TASK_141 = Task(
    id="task_141_grep_classnames",
    name="List class names defined under lib/*.py",
    tags=("grep", "search", "execute", "medium"),
    prompt=(
        "В каталоге lib лежат 6 .py-файлов. Найди в них все определения"
        " классов (строки вида 'class Name:' в начале строки) и собери имена"
        " классов в файл classes.txt — отсортированные в алфавитном порядке"
        " по возрастанию, по одному на строку, только имя класса (без 'class'"
        " и двоеточия)."
    ),
    setup_files=_CLASS_FILES,
    gold_files={"classes.txt": "\n".join(_CLASS_SORTED) + "\n"},
    verifier=file_lines_equal("classes.txt", _CLASS_SORTED),
)


def _make_tests_project() -> tuple[dict[str, str], int]:
    files = {}
    total_asserts = 0
    for i in range(10):
        n_asserts = (i % 3) + 2  # 2..4
        body = f"def test_{i}():\n"
        for j in range(n_asserts):
            body += f"    assert {j} == {j}\n"
        body += "\n"
        files[f"tests/test_{i:02d}.py"] = body
        total_asserts += n_asserts
    return files, total_asserts


_TESTS_FILES, _TESTS_ASSERTS = _make_tests_project()
TASK_142 = Task(
    id="task_142_grep_asserts",
    name="Count `assert ` occurrences in tests/*.py",
    tags=("grep", "search", "execute", "medium"),
    prompt=(
        "В каталоге tests лежат 10 .py-файлов. Посчитай общее число строк,"
        " в которых встречается подстрока 'assert ' (assert с пробелом после),"
        " во всех этих файлах. Запиши число одной строкой в файл assert_count.txt."
    ),
    setup_files=_TESTS_FILES,
    gold_files={"assert_count.txt": f"{_TESTS_ASSERTS}\n"},
    verifier=file_text_equals("assert_count.txt", str(_TESTS_ASSERTS)),
)


def _make_largest_file_project() -> tuple[dict[str, str], str]:
    files = {}
    # File index i has (i+5)*3 + 1 lines; file 09 has the most.
    lengths = {}
    for i in range(10):
        n = (i + 5) * 3 + 1
        files[f"big_project/file_{i:02d}.py"] = "\n".join(f"line {j}" for j in range(n)) + "\n"
        lengths[f"file_{i:02d}.py"] = n
    biggest = max(lengths, key=lambda k: lengths[k])
    return files, biggest


_BIG_FILES, _BIG_BIGGEST = _make_largest_file_project()


def _verify_task_143(ws: Path) -> VerifyResult:
    p = ws / "largest.txt"
    if not p.exists():
        return VerifyResult(False, "largest.txt missing")
    actual = p.read_text().strip()
    expected_variants = {_BIG_BIGGEST, f"big_project/{_BIG_BIGGEST}"}
    if actual in expected_variants:
        return VerifyResult(True, f"largest.txt names the right file: {actual}")
    return VerifyResult(
        False, f"largest.txt is {actual!r}, expected one of {sorted(expected_variants)}"
    )


TASK_143 = Task(
    id="task_143_grep_largest_file",
    name="Find the .py file under big_project/ with the most lines",
    tags=("grep", "compute", "execute", "hard"),
    prompt=(
        "В каталоге big_project лежит 10 .py-файлов разного размера. Найди"
        " среди них файл с наибольшим количеством строк и запиши его имя"
        " (только имя файла, без префикса каталога — либо относительный путь"
        " вида 'big_project/<file>') одной строкой в файл largest.txt."
        " Для подсчёта строк можно использовать 'wc -l' через execute или"
        " любой другой подход."
    ),
    setup_files=_BIG_FILES,
    gold_files={"largest.txt": _BIG_BIGGEST + "\n"},
    verifier=_verify_task_143,
)


def _make_duplicate_funcs_project() -> tuple[dict[str, str], set[str]]:
    files = {
        "src/a.py": "def foo():\n    pass\n\n\ndef bar():\n    pass\n",
        "src/b.py": "def baz():\n    pass\n\n\ndef foo():\n    return 1\n",
        "src/c.py": "def qux():\n    pass\n\n\ndef bar():\n    return 2\n",
        "src/d.py": "def fresh():\n    return 0\n",
    }
    # foo and bar appear in two files; baz, qux, fresh — once.
    return files, {"foo", "bar"}


_DUP_FILES, _DUP_NAMES = _make_duplicate_funcs_project()


def _verify_task_144(ws: Path) -> VerifyResult:
    p = ws / "duplicates.txt"
    if not p.exists():
        return VerifyResult(False, "duplicates.txt missing")
    raw = {line.strip() for line in p.read_text().splitlines() if line.strip()}
    if raw == _DUP_NAMES:
        return VerifyResult(True, "duplicates.txt lists exactly the duplicate function names")
    return VerifyResult(
        False, f"duplicates.txt {sorted(raw)} differs from expected {sorted(_DUP_NAMES)}"
    )


TASK_144 = Task(
    id="task_144_grep_duplicate_funcs",
    name="Find function names defined in 2+ files under src/",
    tags=("grep", "search", "compute", "execute", "hard"),
    prompt=(
        "В каталоге src четыре .py-файла. Найди имена функций (определяемых"
        " как 'def <name>(' в начале строки), которые встречаются в двух и"
        " более файлах. Запиши такие имена в файл duplicates.txt по одному на"
        " строку. Порядок не важен."
    ),
    setup_files=_DUP_FILES,
    gold_files={"duplicates.txt": "\n".join(sorted(_DUP_NAMES)) + "\n"},
    verifier=_verify_task_144,
)


# ---------------------------------------------------------------------------
# Group H: Logs (3 tasks, 145..147)
# ---------------------------------------------------------------------------


def _make_access_log() -> tuple[str, str, int]:
    """Generate a deterministic Apache-like access log.

    Returns (log_text, top_ip, error_count).
    """
    # 100 lines, three IPs with known counts.
    ip_counts = {"10.0.0.1": 40, "10.0.0.2": 35, "10.0.0.3": 25}
    lines = []
    for ip, count in ip_counts.items():
        for i in range(count):
            status = 500 if (ip == "10.0.0.3" and i % 5 == 0) else 200
            lines.append(
                f'{ip} - - [13/May/2026:10:{i % 60:02d}:00 +0000] '
                f'"GET /api/items HTTP/1.1" {status} 123'
            )
    text = "\n".join(lines) + "\n"
    top_ip = max(ip_counts, key=lambda k: ip_counts[k])
    err_count = sum(1 for line in lines if " 500 " in line)
    return text, top_ip, err_count


_ACCESS_LOG, _TOP_IP, _ERR_COUNT = _make_access_log()

TASK_145 = Task(
    id="task_145_log_top_ip",
    name="Find IP with most requests in access.log",
    tags=("logs", "compute", "execute", "hard"),
    prompt=(
        "В файле access.log примерно 100 строк в формате Apache combined log."
        " Первое поле каждой строки — IP-адрес клиента. Найди IP с наибольшим"
        " числом запросов и запиши его одной строкой в файл top_ip.txt (только"
        " сам IP, без других символов)."
    ),
    setup_files={"access.log": _ACCESS_LOG},
    gold_files={"top_ip.txt": _TOP_IP + "\n"},
    verifier=file_text_equals("top_ip.txt", _TOP_IP),
)


TASK_146 = Task(
    id="task_146_log_count_5xx",
    name="Count 5xx responses in access.log",
    tags=("logs", "compute", "execute", "medium"),
    prompt=(
        "В файле access.log примерно 100 строк в формате Apache combined log."
        " Седьмой токен в каждой строке — это HTTP-статус (целое число от 100"
        " до 599). Посчитай количество строк, у которых статус начинается с"
        " цифры 5 (то есть 500..599 — серверные ошибки), и запиши число одной"
        " строкой в файл err_count.txt."
    ),
    setup_files={"access.log": _ACCESS_LOG},
    gold_files={"err_count.txt": f"{_ERR_COUNT}\n"},
    verifier=file_text_equals("err_count.txt", str(_ERR_COUNT)),
)


# 147. Filter access.log to a specific status
def _make_status_log() -> tuple[str, str]:
    lines = [
        '10.0.0.1 - - [13/May/2026:10:00:00 +0000] "GET / HTTP/1.1" 200 100',
        '10.0.0.2 - - [13/May/2026:10:00:01 +0000] "GET /a HTTP/1.1" 404 50',
        '10.0.0.1 - - [13/May/2026:10:00:02 +0000] "GET /b HTTP/1.1" 200 200',
        '10.0.0.3 - - [13/May/2026:10:00:03 +0000] "POST /x HTTP/1.1" 500 0',
        '10.0.0.2 - - [13/May/2026:10:00:04 +0000] "GET /c HTTP/1.1" 404 30',
        '10.0.0.1 - - [13/May/2026:10:00:05 +0000] "GET /d HTTP/1.1" 200 70',
        '10.0.0.4 - - [13/May/2026:10:00:06 +0000] "GET /e HTTP/1.1" 404 25',
    ]
    full = "\n".join(lines) + "\n"
    gold = "\n".join(line for line in lines if " 404 " in line) + "\n"
    return full, gold


_STATUS_LOG, _STATUS_GOLD = _make_status_log()


def _verify_task_147(ws: Path) -> VerifyResult:
    p = ws / "not_found.log"
    if not p.exists():
        return VerifyResult(False, "not_found.log missing")
    actual = p.read_text().strip()
    expected = _STATUS_GOLD.strip()
    if actual == expected:
        return VerifyResult(True, "not_found.log has the three 404 lines")
    return VerifyResult(False, f"not_found.log content differs\n--- got ---\n{actual}\n--- exp ---\n{expected}")


TASK_147 = Task(
    id="task_147_log_filter_404",
    name="Filter access.log to 404 lines",
    tags=("logs", "filter", "execute", "medium"),
    prompt=(
        "В файле access.log семь строк в формате Apache combined log."
        " Сохрани в файл not_found.log только те строки, у которых HTTP-статус"
        " равен 404, в исходном порядке. Содержимое каждой строки переноси"
        " байт-в-байт."
    ),
    setup_files={"access.log": _STATUS_LOG},
    gold_files={"not_found.log": _STATUS_GOLD},
    verifier=_verify_task_147,
)


# ---------------------------------------------------------------------------
# Group I: Mixed / hard (3 tasks, 148..150)
# ---------------------------------------------------------------------------


# 148. Convert csv to xlsx
def _xlsx_148_setup(_ws: Path) -> None:
    # No setup callback needed — input is plain CSV. Kept for symmetry.
    return None


def _xlsx_148_gold_callback(ws: Path) -> None:
    import openpyxl  # noqa: PLC0415

    wb = openpyxl.Workbook()
    sh = wb.active
    sh.title = "Sheet1"
    sh.append(["name", "score"])
    sh.append(["Alice", 90])
    sh.append(["Bob", 75])
    sh.append(["Carol", 100])
    wb.save(ws / "scores.xlsx")


def _verify_task_148(ws: Path) -> VerifyResult:
    import openpyxl  # noqa: PLC0415

    p = ws / "scores.xlsx"
    if not p.exists():
        return VerifyResult(False, "scores.xlsx missing")
    wb = openpyxl.load_workbook(p, data_only=True)
    sh = wb[wb.sheetnames[0]]
    rows = [tuple(c.value for c in row) for row in sh.iter_rows(values_only=False)]
    expected = [("name", "score"), ("Alice", 90), ("Bob", 75), ("Carol", 100)]
    if len(rows) != len(expected):
        return VerifyResult(False, f"scores.xlsx has {len(rows)} rows, expected {len(expected)}")
    # Row 0 is the header — compare strings; data rows compare name as string,
    # score as int (so we accept openpyxl reading 90 as int or float 90.0).
    if rows[0] != expected[0]:
        return VerifyResult(False, f"header row mismatch: {rows[0]!r} vs {expected[0]!r}")
    for got, exp in zip(rows[1:], expected[1:], strict=False):
        if got[0] != exp[0]:
            return VerifyResult(False, f"row name mismatch: {got!r} vs {exp!r}")
        try:
            if int(got[1]) != int(exp[1]):
                return VerifyResult(False, f"row score mismatch: {got!r} vs {exp!r}")
        except (TypeError, ValueError):
            return VerifyResult(False, f"row score not numeric: {got!r}")
    return VerifyResult(True, "scores.xlsx contains the three rows from CSV")


TASK_148 = Task(
    id="task_148_csv_to_xlsx",
    name="Convert scores.csv to scores.xlsx",
    tags=("csv", "xlsx", "convert", "execute", "hard"),
    prompt=(
        "В файле scores.csv колонки 'name,score' и три строки данных (Alice,90"
        "; Bob,75; Carol,100). Преобразуй его в Excel-файл scores.xlsx (на"
        " активном листе должны быть те же 4 строки: заголовок + три строки"
        " данных в исходном порядке; значения score сохрани как числа)."
        " Используй openpyxl — он установлен в окружении."
    ),
    setup_files={"scores.csv": "name,score\nAlice,90\nBob,75\nCarol,100\n"},
    setup_callback=_xlsx_148_setup,
    gold_callback=_xlsx_148_gold_callback,
    verifier=_verify_task_148,
)


# 149. SQLite + JSON export
def _sqlite_149_setup(ws: Path) -> None:
    conn = sqlite3.connect(ws / "products.db")
    conn.execute("CREATE TABLE products (id INTEGER PRIMARY KEY, name TEXT, price INTEGER)")
    rows = [(1, "Apple", 10), (2, "Bread", 25), (3, "Cheese", 80), (4, "Donut", 5)]
    conn.executemany("INSERT INTO products VALUES (?, ?, ?)", rows)
    conn.commit()
    conn.close()


_PRODUCTS_GOLD = [
    {"id": 1, "name": "Apple", "price": 10},
    {"id": 2, "name": "Bread", "price": 25},
    {"id": 3, "name": "Cheese", "price": 80},
    {"id": 4, "name": "Donut", "price": 5},
]


def _verify_task_149(ws: Path) -> VerifyResult:
    p = ws / "products.json"
    if not p.exists():
        return VerifyResult(False, "products.json missing")
    try:
        data = json.loads(p.read_text())
    except json.JSONDecodeError as exc:
        return VerifyResult(False, f"products.json invalid JSON: {exc}")
    if not isinstance(data, list):
        return VerifyResult(False, f"products.json must be a list, got {type(data).__name__}")
    if len(data) != len(_PRODUCTS_GOLD):
        return VerifyResult(False, f"products.json has {len(data)} records, expected {len(_PRODUCTS_GOLD)}")
    # Order-insensitive by id.
    actual_by_id = {row.get("id"): row for row in data}
    for expected in _PRODUCTS_GOLD:
        row = actual_by_id.get(expected["id"])
        if row is None:
            return VerifyResult(False, f"id {expected['id']!r} missing from products.json")
        if row.get("name") != expected["name"]:
            return VerifyResult(False, f"id {expected['id']}: name mismatch {row.get('name')!r}")
        try:
            if int(row.get("price")) != expected["price"]:
                return VerifyResult(False, f"id {expected['id']}: price mismatch {row.get('price')!r}")
        except (TypeError, ValueError):
            return VerifyResult(False, f"id {expected['id']}: price not numeric: {row.get('price')!r}")
    return VerifyResult(True, "products.json has all 4 products with correct fields")


TASK_149 = Task(
    id="task_149_sqlite_to_json",
    name="Export products table from SQLite to JSON",
    tags=("sqlite", "json", "convert", "execute", "hard"),
    prompt=(
        "В корне рабочей директории есть SQLite-база products.db с таблицей"
        " products (колонки id INTEGER, name TEXT, price INTEGER). Прочитай"
        " все строки и сохрани их в файл products.json — как валидный"
        " JSON-массив из объектов, у каждого должны быть ключи id, name,"
        " price (price — число). Порядок объектов в массиве не важен."
    ),
    setup_files={},
    setup_callback=_sqlite_149_setup,
    gold_files={"products.json": json.dumps(_PRODUCTS_GOLD) + "\n"},
    verifier=_verify_task_149,
)


# 150. Python script that reads csv, sums column, runs, and writes result
_TX150_AMOUNTS = [10, 25, 17, 33, 48, 19, 22, 41, 8, 67]
_TX150_TOTAL = sum(_TX150_AMOUNTS)  # 290
_TX150_CSV = "id,amount\n" + "".join(f"{i + 1},{a}\n" for i, a in enumerate(_TX150_AMOUNTS))


def _verify_task_150(ws: Path) -> VerifyResult:
    if not (ws / "sum.py").exists():
        return VerifyResult(False, "sum.py missing")
    if not (ws / "total.txt").exists():
        return VerifyResult(False, "total.txt missing")
    actual = (ws / "total.txt").read_text().strip()
    if actual != str(_TX150_TOTAL):
        return VerifyResult(False, f"total.txt is {actual!r}, expected {_TX150_TOTAL!r}")
    return VerifyResult(True, f"sum.py exists and total.txt == {_TX150_TOTAL}")


TASK_150 = Task(
    id="task_150_script_csv_total",
    name="Write sum.py + run it to produce total.txt",
    tags=("python", "csv", "execute", "hard"),
    prompt=(
        "В файле transactions.csv два столбца: id,amount, плюс заголовок (10"
        " строк данных). Сделай две вещи:\n"
        "  1) создай файл sum.py с программой на Python, которая читает"
        " transactions.csv, суммирует значения столбца amount и печатает"
        " получившееся число (без других символов) в стандартный вывод;\n"
        "  2) запусти этот скрипт (через execute) и сохрани его вывод в файл"
        " total.txt. В total.txt должна оказаться одна строка с числом."
    ),
    setup_files={"transactions.csv": _TX150_CSV},
    gold_files={
        "sum.py": (
            "import csv\n"
            "\n"
            "\n"
            "total = 0\n"
            "with open('transactions.csv') as f:\n"
            "    reader = csv.DictReader(f)\n"
            "    for row in reader:\n"
            "        total += int(row['amount'])\n"
            "print(total)\n"
        ),
        "total.txt": f"{_TX150_TOTAL}\n",
    },
    verifier=_verify_task_150,
)


HARD_TASKS: list[Task] = [
    TASK_101, TASK_102, TASK_103, TASK_104, TASK_105,
    TASK_106, TASK_107, TASK_108, TASK_109, TASK_110,
    TASK_111, TASK_112, TASK_113,
    TASK_114, TASK_115, TASK_116, TASK_117, TASK_118,
    TASK_119, TASK_120, TASK_121, TASK_122,
    TASK_123, TASK_124,
    TASK_125, TASK_126, TASK_127, TASK_128, TASK_129,
    TASK_130, TASK_131, TASK_132, TASK_133, TASK_134,
    TASK_135, TASK_136, TASK_137, TASK_138, TASK_139,
    TASK_140, TASK_141, TASK_142, TASK_143, TASK_144,
    TASK_145, TASK_146, TASK_147,
    TASK_148, TASK_149, TASK_150,
]


# Suppress "imported but unused" warnings for helpers kept around for future
# tasks (file_contains / file_matches_regex are used in some current task
# bodies but ruff doesn't always notice through `all_of`).
_ = (csv, io, re, file_contains, file_matches_regex)
