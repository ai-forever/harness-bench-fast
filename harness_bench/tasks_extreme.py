"""Tasks 151..200 — extreme benchmark wave.

This module ramps up complexity beyond tasks 101..150:
- Composite pipelines that chain CSV/SQLite/XLSX/JSONL together.
- Archive operations (zip, gzip, tar) — binary formats requiring execute.
- Multi-file refactoring across an entire mock project tree.
- Algorithms with pytest checks (quicksort, LRU cache, linked list, trees).
- Real data-analysis aggregates: rolling averages, histograms, z-scores,
  pivot tables, percentiles.
- XML/markdown parsing and cross-format conversion.
- Hard composites: 3-way joins, hourly log aggregation, top-N per group.

All tasks ship a deterministic setup + gold pair so `verify-gold` exercises
the verifier without spending any LLM tokens.
"""

from __future__ import annotations

import gzip
import io
import json
import math
import shutil
import sqlite3
import statistics
import tarfile
import zipfile
from collections import Counter, defaultdict
from pathlib import Path

import yaml

from harness_bench.core import Task, VerifyResult
from harness_bench.verifiers import (
    all_of,
    file_contains,
    file_exists,
    file_lines_equal,
    file_matches_regex,
    file_not_contains,
    file_text_equals,
    pytest_passes,
    python_callable_returns,
    sqlite_query_returns,
    xlsx_cell_equals,
)

# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------


def _json_file_matches_loose(rel: str, expected, *, ordered: bool = False):
    """Verifier: load `rel` as JSON, compare to `expected`.

    When `ordered=False` and both values are lists of dicts, compare them
    as sets-of-tuples (i.e. order-insensitive). Otherwise compare directly.
    """

    def _canon(value):
        if isinstance(value, list):
            return [_canon(v) for v in value]
        if isinstance(value, dict):
            return {k: _canon(v) for k, v in value.items()}
        return value

    def _normalise(value):
        if not ordered and isinstance(value, list) and all(isinstance(v, dict) for v in value):
            return sorted([_canon(v) for v in value], key=lambda d: json.dumps(d, sort_keys=True))
        return _canon(value)

    def _check(ws: Path) -> VerifyResult:
        p = ws / rel
        if not p.exists():
            return VerifyResult(False, f"{rel} missing")
        try:
            data = json.loads(p.read_text())
        except json.JSONDecodeError as exc:
            return VerifyResult(False, f"{rel} invalid JSON: {exc}")
        if _normalise(data) == _normalise(expected):
            return VerifyResult(True, f"{rel} matches expected JSON")
        return VerifyResult(False, f"{rel} JSON mismatch\n--- got ---\n{data}\n--- exp ---\n{expected}")

    return _check


# ---------------------------------------------------------------------------
# Group A: Composite pipelines (8 tasks, 151..158)
# ---------------------------------------------------------------------------

# 151. csv_filter_groupby_to_json
_ORDERS_ROWS = [
    ("alice", "books", 120), ("bob", "tech", 95), ("carol", "books", 250),
    ("dave", "food", 80), ("eve", "tech", 1100), ("frank", "books", 40),
    ("grace", "food", 110), ("henry", "tech", 105), ("ivy", "books", 300),
    ("jack", "food", 200), ("kate", "tech", 99), ("liam", "books", 175),
    ("mia", "food", 130), ("noah", "tech", 1500), ("olive", "books", 95),
    ("paul", "food", 175), ("quinn", "tech", 80), ("riley", "books", 220),
    ("sam", "food", 60), ("tara", "tech", 150),
]
_ORDERS_CSV = "user,category,amount\n" + "".join(f"{u},{c},{a}\n" for u, c, a in _ORDERS_ROWS)
# Filter amount >= 100, group by category, sum, sort desc by sum:
_filtered = [(c, a) for _u, c, a in _ORDERS_ROWS if a >= 100]
_sums: dict[str, int] = {}
for c, a in _filtered:
    _sums[c] = _sums.get(c, 0) + a
_ORDERS_GOLD = [
    {"category": c, "total": s}
    for c, s in sorted(_sums.items(), key=lambda kv: (-kv[1], kv[0]))
]
TASK_151 = Task(
    id="task_151_pipeline_csv_to_json",
    name="Filter + groupby + sort orders.csv to summary.json",
    tags=("csv", "json", "pipeline", "execute", "hard"),
    prompt=(
        "В файле orders.csv 20 строк данных (user,category,amount). Сделай"
        " summary.json — JSON-массив объектов с полями category и total."
        " Включай только заказы, у которых amount >= 100; суммируй amount по"
        " каждой категории; сортируй итоговый массив по полю total по убыванию"
        " (при равенстве — по имени категории по возрастанию). В итоговом"
        " массиве должны быть только категории, по которым есть хотя бы одна"
        " подходящая строка."
    ),
    setup_files={"orders.csv": _ORDERS_CSV},
    gold_files={"summary.json": json.dumps(_ORDERS_GOLD) + "\n"},
    verifier=_json_file_matches_loose("summary.json", _ORDERS_GOLD, ordered=True),
)


# 152. sqlite_join_to_json
def _sqlite_152_setup(ws: Path) -> None:
    conn = sqlite3.connect(ws / "shop.db")
    conn.executescript(
        """
        CREATE TABLE users (id INTEGER PRIMARY KEY, name TEXT);
        CREATE TABLE orders (id INTEGER PRIMARY KEY, user_id INTEGER, amount INTEGER);
        INSERT INTO users VALUES (1, 'Alice'), (2, 'Bob'), (3, 'Carol'), (4, 'Dave');
        INSERT INTO orders (user_id, amount) VALUES
            (1, 100), (1, 50), (2, 200), (3, 300), (3, 80), (4, 70), (1, 90);
        """
    )
    conn.commit()
    conn.close()


_SQLITE_152_GOLD = [
    {"name": "Alice", "total": 240},
    {"name": "Bob", "total": 200},
    {"name": "Carol", "total": 380},
    {"name": "Dave", "total": 70},
]
TASK_152 = Task(
    id="task_152_sqlite_join_to_json",
    name="JOIN users+orders in sqlite, export per-user totals to JSON",
    tags=("sqlite", "json", "pipeline", "execute", "hard"),
    prompt=(
        "В файле shop.db лежат таблицы users(id,name) и orders(id,user_id,amount)."
        " Для каждого пользователя посчитай сумму всех его заказов и сохрани"
        " результат в users_totals.json как JSON-массив объектов с полями"
        " name и total (целое число). Включай всех 4 пользователей (даже"
        " если у кого-то один заказ). Порядок объектов в массиве не важен."
    ),
    setup_files={},
    setup_callback=_sqlite_152_setup,
    gold_files={"users_totals.json": json.dumps(_SQLITE_152_GOLD) + "\n"},
    verifier=_json_file_matches_loose("users_totals.json", _SQLITE_152_GOLD),
)


# 153. xlsx_summary_md — read xlsx, compute totals per region, write markdown report
def _xlsx_153_setup(ws: Path) -> None:
    import openpyxl  # noqa: PLC0415

    wb = openpyxl.Workbook()
    sh = wb.active
    sh.title = "Sales"
    sh.append(["region", "amount"])
    rows = [
        ("EU", 100), ("US", 200), ("APAC", 50), ("EU", 75), ("US", 125),
        ("APAC", 25), ("EU", 50), ("US", 175), ("APAC", 100),
    ]
    for r in rows:
        sh.append(list(r))
    wb.save(ws / "sales.xlsx")


_XLSX_153_TOTALS = {"APAC": 175, "EU": 225, "US": 500}
_XLSX_153_GOLD_MD = (
    "| region | total |\n"
    "| --- | --- |\n"
    + "".join(f"| {r} | {t} |\n" for r, t in sorted(_XLSX_153_TOTALS.items()))
)


def _verify_task_153(ws: Path) -> VerifyResult:
    p = ws / "report.md"
    if not p.exists():
        return VerifyResult(False, "report.md missing")
    text = p.read_text()
    required = ["| region", "| total", "---", "EU", "US", "APAC", "225", "500", "175"]
    missing = [s for s in required if s not in text]
    if missing:
        return VerifyResult(False, f"report.md missing pieces: {missing!r}")
    return VerifyResult(True, "report.md has the three regions with correct totals")


TASK_153 = Task(
    id="task_153_xlsx_to_markdown_report",
    name="Read sales.xlsx, write markdown report.md",
    tags=("xlsx", "markdown", "pipeline", "execute", "hard"),
    prompt=(
        "В файле sales.xlsx (лист Sales) колонки region и amount, 9 строк"
        " данных. Посчитай сумму amount по каждому региону и оформи результат"
        " в виде markdown-таблицы в файле report.md. Заголовок таблицы —"
        " '| region | total |', под ним строка-разделитель '| --- | --- |',"
        " дальше — по строке на каждый регион (EU, US, APAC), значение —"
        " посчитанная сумма. Порядок регионов в отчёте не важен."
    ),
    setup_files={},
    setup_callback=_xlsx_153_setup,
    gold_files={"report.md": _XLSX_153_GOLD_MD},
    verifier=_verify_task_153,
)


# 154. jsonl_into_sqlite
_EVENTS_154 = [
    {"id": i + 1, "kind": ["click", "view", "buy"][i % 3], "value": (i + 1) * 10}
    for i in range(15)
]
_JSONL_154 = "".join(json.dumps(e) + "\n" for e in _EVENTS_154)
_SQLITE_154_GOLD_COUNT = len(_EVENTS_154)


def _verify_task_154(ws: Path) -> VerifyResult:
    p = ws / "events.db"
    if not p.exists():
        return VerifyResult(False, "events.db missing")
    try:
        conn = sqlite3.connect(p)
        rows = conn.execute("SELECT id, kind, value FROM events ORDER BY id").fetchall()
        conn.close()
    except sqlite3.Error as exc:
        return VerifyResult(False, f"events.db sqlite error: {exc}")
    if len(rows) != _SQLITE_154_GOLD_COUNT:
        return VerifyResult(False, f"events table has {len(rows)} rows, expected {_SQLITE_154_GOLD_COUNT}")
    for i, row in enumerate(rows):
        expected = _EVENTS_154[i]
        if (row[0], row[1], row[2]) != (expected["id"], expected["kind"], expected["value"]):
            return VerifyResult(False, f"row {i}: got {row!r}, expected {expected!r}")
    return VerifyResult(True, "events.db contains all 15 records from events.jsonl")


def _sqlite_154_gold_callback(ws: Path) -> None:
    conn = sqlite3.connect(ws / "events.db")
    conn.execute("CREATE TABLE events (id INTEGER PRIMARY KEY, kind TEXT, value INTEGER)")
    conn.executemany(
        "INSERT INTO events VALUES (?, ?, ?)",
        [(e["id"], e["kind"], e["value"]) for e in _EVENTS_154],
    )
    conn.commit()
    conn.close()


TASK_154 = Task(
    id="task_154_jsonl_into_sqlite",
    name="Load events.jsonl into a fresh sqlite events.db",
    tags=("jsonl", "sqlite", "pipeline", "execute", "hard"),
    prompt=(
        "В файле events.jsonl 15 JSON-объектов (по одному на строку) с полями"
        " id, kind, value. Создай SQLite-базу events.db, в ней таблицу events"
        " с колонками id INTEGER PRIMARY KEY, kind TEXT, value INTEGER, и"
        " вставь туда все 15 записей из events.jsonl. Используй стандартный"
        " модуль sqlite3."
    ),
    setup_files={"events.jsonl": _JSONL_154},
    gold_callback=_sqlite_154_gold_callback,
    verifier=_verify_task_154,
)


# 155. log_count_per_endpoint
def _make_access_log_155() -> tuple[str, dict[str, int]]:
    endpoints = {"/api/users": 12, "/api/items": 18, "/api/health": 5, "/login": 9}
    lines = []
    for ep, count in endpoints.items():
        for _ in range(count):
            lines.append(f'10.0.0.1 - - [13/May/2026:10:00:00 +0000] "GET {ep} HTTP/1.1" 200 100')
    return "\n".join(lines) + "\n", endpoints


_ACCESS_155, _ENDPOINT_COUNTS_155 = _make_access_log_155()
_ENDPOINT_CSV_155 = "endpoint,count\n" + "".join(
    f"{ep},{cnt}\n" for ep, cnt in sorted(_ENDPOINT_COUNTS_155.items())
)


def _verify_task_155(ws: Path) -> VerifyResult:
    p = ws / "endpoints.csv"
    if not p.exists():
        return VerifyResult(False, "endpoints.csv missing")
    import csv as _csv  # noqa: PLC0415

    rows = list(_csv.DictReader(io.StringIO(p.read_text())))
    if not rows:
        return VerifyResult(False, "endpoints.csv empty (no data rows)")
    actual = {r["endpoint"]: int(r["count"]) for r in rows}
    if actual == _ENDPOINT_COUNTS_155:
        return VerifyResult(True, "endpoints.csv counts match")
    return VerifyResult(False, f"endpoints.csv counts {actual!r} differ from {_ENDPOINT_COUNTS_155!r}")


TASK_155 = Task(
    id="task_155_log_count_per_endpoint",
    name="Count log requests per endpoint, write endpoints.csv",
    tags=("logs", "csv", "pipeline", "execute", "hard"),
    prompt=(
        "В файле access.log около 44 строк в формате Apache combined log."
        " Седьмой токен в каждой строке — путь URL (например, /api/users)."
        " Посчитай количество запросов по каждому уникальному endpoint и"
        " сохрани результат в endpoints.csv с заголовком 'endpoint,count' —"
        " по одной строке на endpoint. Порядок строк не важен."
    ),
    setup_files={"access.log": _ACCESS_155},
    gold_files={"endpoints.csv": _ENDPOINT_CSV_155},
    verifier=_verify_task_155,
)


# 156. csv_normalize_split — valid rows -> good.csv, invalid -> bad.csv
_NORM_ROWS = [
    ("u01", "ok", 100), ("u02", "BAD", -5), ("u03", "ok", 50),
    ("u04", "ok", 0), ("u05", "BAD", 200), ("u06", "ok", 80),
    ("u07", "ok", -1), ("u08", "ok", 999), ("u09", "BAD", 10),
    ("u10", "ok", 45),
]
_NORM_CSV = "id,status,value\n" + "".join(f"{i},{s},{v}\n" for i, s, v in _NORM_ROWS)
_GOOD_ROWS = [(i, s, v) for i, s, v in _NORM_ROWS if s == "ok" and v >= 0]
_BAD_ROWS = [(i, s, v) for i, s, v in _NORM_ROWS if not (s == "ok" and v >= 0)]
_GOOD_CSV = "id,status,value\n" + "".join(f"{i},{s},{v}\n" for i, s, v in _GOOD_ROWS)
_BAD_CSV = "id,status,value\n" + "".join(f"{i},{s},{v}\n" for i, s, v in _BAD_ROWS)
TASK_156 = Task(
    id="task_156_csv_validate_split",
    name="Split records.csv into good.csv and bad.csv",
    tags=("csv", "filter", "pipeline", "execute", "hard"),
    prompt=(
        "В файле records.csv 10 строк данных (id,status,value). Раздели их"
        " на два файла:\n"
        "  - good.csv — те, у которых status == 'ok' И value >= 0;\n"
        "  - bad.csv — все остальные (status != 'ok' или value < 0).\n"
        "В обоих файлах должна быть исходная строка-заголовок 'id,status,value',"
        " а данные — в исходном порядке."
    ),
    setup_files={"records.csv": _NORM_CSV},
    gold_files={"good.csv": _GOOD_CSV, "bad.csv": _BAD_CSV},
    verifier=all_of(
        file_text_equals("good.csv", _GOOD_CSV),
        file_text_equals("bad.csv", _BAD_CSV),
    ),
)


# 157. two_csv_diff — report rows only in A, only in B
_A_ROWS = {"alpha", "bravo", "charlie", "delta", "echo"}
_B_ROWS = {"bravo", "charlie", "foxtrot", "golf", "echo"}
_A_CSV = "name\n" + "\n".join(_A_ROWS) + "\n"
_B_CSV = "name\n" + "\n".join(_B_ROWS) + "\n"
_ONLY_A = sorted(_A_ROWS - _B_ROWS)  # alpha, delta
_ONLY_B = sorted(_B_ROWS - _A_ROWS)  # foxtrot, golf


def _verify_task_157(ws: Path) -> VerifyResult:
    p1 = ws / "only_in_a.txt"
    p2 = ws / "only_in_b.txt"
    if not p1.exists() or not p2.exists():
        return VerifyResult(False, "only_in_a.txt or only_in_b.txt missing")
    a = sorted(line.strip() for line in p1.read_text().splitlines() if line.strip())
    b = sorted(line.strip() for line in p2.read_text().splitlines() if line.strip())
    if a != _ONLY_A:
        return VerifyResult(False, f"only_in_a.txt {a!r} differs from {_ONLY_A!r}")
    if b != _ONLY_B:
        return VerifyResult(False, f"only_in_b.txt {b!r} differs from {_ONLY_B!r}")
    return VerifyResult(True, "diff files have the expected names")


TASK_157 = Task(
    id="task_157_two_csv_diff",
    name="Diff two name lists into only_in_a.txt and only_in_b.txt",
    tags=("csv", "diff", "pipeline", "execute", "medium"),
    prompt=(
        "В файлах a.csv и b.csv по одному столбцу 'name'. Создай два файла:\n"
        "  - only_in_a.txt — имена, которые есть в a.csv, но не в b.csv;\n"
        "  - only_in_b.txt — имена, которые есть в b.csv, но не в a.csv.\n"
        "В каждом файле по одному имени на строку, порядок не важен."
    ),
    setup_files={"a.csv": _A_CSV, "b.csv": _B_CSV},
    gold_files={
        "only_in_a.txt": "\n".join(_ONLY_A) + "\n",
        "only_in_b.txt": "\n".join(_ONLY_B) + "\n",
    },
    verifier=_verify_task_157,
)


# 158. xlsx_split_sheets — workbook with 3 sheets -> 3 CSV files
def _xlsx_158_setup(ws: Path) -> None:
    import openpyxl  # noqa: PLC0415

    wb = openpyxl.Workbook()
    wb.active.title = "Q1"
    wb["Q1"].append(["month", "sales"])
    for m, s in [("Jan", 100), ("Feb", 150), ("Mar", 200)]:
        wb["Q1"].append([m, s])
    q2 = wb.create_sheet("Q2")
    q2.append(["month", "sales"])
    for m, s in [("Apr", 250), ("May", 175), ("Jun", 300)]:
        q2.append([m, s])
    q3 = wb.create_sheet("Q3")
    q3.append(["month", "sales"])
    for m, s in [("Jul", 350), ("Aug", 400), ("Sep", 325)]:
        q3.append([m, s])
    wb.save(ws / "quarterly.xlsx")


_XLSX_158_SHEETS = {
    "Q1.csv": "month,sales\nJan,100\nFeb,150\nMar,200\n",
    "Q2.csv": "month,sales\nApr,250\nMay,175\nJun,300\n",
    "Q3.csv": "month,sales\nJul,350\nAug,400\nSep,325\n",
}


def _verify_task_158(ws: Path) -> VerifyResult:
    for fname, expected in _XLSX_158_SHEETS.items():
        p = ws / fname
        if not p.exists():
            return VerifyResult(False, f"{fname} missing")
        actual = p.read_text().strip()
        if actual != expected.strip():
            return VerifyResult(False, f"{fname} content differs from expected")
    return VerifyResult(True, "all three CSVs match the corresponding xlsx sheets")


TASK_158 = Task(
    id="task_158_xlsx_split_sheets",
    name="Split quarterly.xlsx into Q1.csv, Q2.csv, Q3.csv",
    tags=("xlsx", "csv", "pipeline", "execute", "hard"),
    prompt=(
        "В файле quarterly.xlsx три листа: Q1, Q2, Q3. У каждого листа одна и"
        " та же схема — колонки month и sales. Для каждого листа создай"
        " отдельный CSV-файл с тем же именем в корне рабочей директории:"
        " Q1.csv, Q2.csv, Q3.csv. В каждом — заголовок 'month,sales' и"
        " соответствующие три строки данных в исходном порядке."
    ),
    setup_files={},
    setup_callback=_xlsx_158_setup,
    gold_files=_XLSX_158_SHEETS,
    verifier=_verify_task_158,
)


# ---------------------------------------------------------------------------
# Group B: Archives (3 tasks, 159..161)
# ---------------------------------------------------------------------------


def _zip_159_setup(ws: Path) -> None:
    archive = ws / "archive.zip"
    with zipfile.ZipFile(archive, "w") as zf:
        zf.writestr("data/alpha.txt", "alpha content\n")
        zf.writestr("data/beta.txt", "beta content\n")
        zf.writestr("data/gamma.txt", "gamma content\n")


def _verify_task_159(ws: Path) -> VerifyResult:
    for name, content in [
        ("alpha.txt", "alpha content"),
        ("beta.txt", "beta content"),
        ("gamma.txt", "gamma content"),
    ]:
        # Accept either flat (extracted/alpha.txt) or preserving the inner
        # `data/` prefix (extracted/data/alpha.txt).
        candidates = [ws / "extracted" / name, ws / "extracted" / "data" / name]
        match = next((p for p in candidates if p.exists()), None)
        if match is None:
            return VerifyResult(False, f"extracted/{name} (or extracted/data/{name}) missing")
        if match.read_text().strip() != content:
            return VerifyResult(False, f"{match}: content {match.read_text()!r} differs from {content!r}")
    return VerifyResult(True, "archive.zip extracted to extracted/ with all three files")


def _zip_159_gold_callback(ws: Path) -> None:
    (ws / "extracted").mkdir(exist_ok=True)
    (ws / "extracted" / "alpha.txt").write_text("alpha content\n")
    (ws / "extracted" / "beta.txt").write_text("beta content\n")
    (ws / "extracted" / "gamma.txt").write_text("gamma content\n")


TASK_159 = Task(
    id="task_159_unzip_extract",
    name="Extract archive.zip into the extracted/ directory",
    tags=("archive", "zip", "execute", "hard"),
    prompt=(
        "В корне рабочей директории есть архив archive.zip. Распакуй его так,"
        " чтобы содержащиеся в нём текстовые файлы оказались в каталоге"
        " extracted/. Внутри архива файлы лежат под путём data/<имя>; куда"
        " положить их при распаковке — в extracted/<имя> или"
        " extracted/data/<имя> — на твоё усмотрение, важно, чтобы можно было"
        " найти их по имени файла."
    ),
    setup_files={},
    setup_callback=_zip_159_setup,
    gold_callback=_zip_159_gold_callback,
    verifier=_verify_task_159,
)


def _zip_160_setup(ws: Path) -> None:
    (ws / "files").mkdir(exist_ok=True)
    (ws / "files" / "one.txt").write_text("one\n")
    (ws / "files" / "two.txt").write_text("two\n")
    (ws / "files" / "three.txt").write_text("three\n")


def _verify_task_160(ws: Path) -> VerifyResult:
    p = ws / "bundle.zip"
    if not p.exists():
        return VerifyResult(False, "bundle.zip missing")
    try:
        with zipfile.ZipFile(p) as zf:
            names = {name.split("/")[-1] for name in zf.namelist() if name.split("/")[-1]}
            expected = {"one.txt", "two.txt", "three.txt"}
            if not expected.issubset(names):
                return VerifyResult(False, f"bundle.zip names {sorted(names)} missing some of {sorted(expected)}")
            for n, want in [("one.txt", "one"), ("two.txt", "two"), ("three.txt", "three")]:
                full = next(name for name in zf.namelist() if name.endswith(n))
                if zf.read(full).decode().strip() != want:
                    return VerifyResult(False, f"{n} content in zip differs from expected")
    except zipfile.BadZipFile:
        return VerifyResult(False, "bundle.zip is not a valid zip")
    return VerifyResult(True, "bundle.zip contains the three files with correct content")


def _zip_160_gold_callback(ws: Path) -> None:
    with zipfile.ZipFile(ws / "bundle.zip", "w") as zf:
        zf.write(ws / "files" / "one.txt", "one.txt")
        zf.write(ws / "files" / "two.txt", "two.txt")
        zf.write(ws / "files" / "three.txt", "three.txt")


TASK_160 = Task(
    id="task_160_create_zip",
    name="Create bundle.zip containing files/one.txt, two.txt, three.txt",
    tags=("archive", "zip", "execute", "medium"),
    prompt=(
        "В каталоге files три файла: one.txt, two.txt, three.txt. Упакуй все"
        " три в архив bundle.zip в корне рабочей директории. Допустимо как"
        " положить файлы под путём files/<имя>, так и просто в корень архива"
        " под именем <имя> — важно, чтобы файлы можно было прочитать из"
        " архива и содержимое каждого совпадало с исходным."
    ),
    setup_files={},
    setup_callback=_zip_160_setup,
    gold_callback=_zip_160_gold_callback,
    verifier=_verify_task_160,
)


def _gzip_161_setup(ws: Path) -> None:
    payload = b"data line\n" * 50
    (ws / "input.txt").write_bytes(payload)


def _verify_task_161(ws: Path) -> VerifyResult:
    p = ws / "input.txt.gz"
    if not p.exists():
        return VerifyResult(False, "input.txt.gz missing")
    try:
        with gzip.open(p, "rb") as fh:
            content = fh.read()
    except (OSError, EOFError) as exc:
        return VerifyResult(False, f"input.txt.gz cannot be decompressed: {exc}")
    expected = b"data line\n" * 50
    if content == expected:
        return VerifyResult(True, "input.txt.gz decompresses to the original input.txt content")
    return VerifyResult(False, f"input.txt.gz decompressed content differs (len={len(content)}, expected {len(expected)})")


def _gzip_161_gold_callback(ws: Path) -> None:
    with open(ws / "input.txt", "rb") as src, gzip.open(ws / "input.txt.gz", "wb") as dst:
        shutil.copyfileobj(src, dst)


TASK_161 = Task(
    id="task_161_gzip_compress",
    name="Gzip-compress input.txt to input.txt.gz",
    tags=("archive", "gzip", "execute", "medium"),
    prompt=(
        "В корне рабочей директории есть текстовый файл input.txt. Создай"
        " рядом сжатый gzip-файл input.txt.gz, чтобы при распаковке (через"
        " gzip.open или gunzip) получалось ровно то же содержимое, что в"
        " input.txt. Сам input.txt не трогай."
    ),
    setup_files={},
    setup_callback=_gzip_161_setup,
    gold_callback=_gzip_161_gold_callback,
    verifier=_verify_task_161,
)


# ---------------------------------------------------------------------------
# Group C: Refactoring (5 tasks, 162..166)
# ---------------------------------------------------------------------------


_BIG_MODULE_SRC = (
    "class Alpha:\n"
    "    def hello(self):\n"
    "        return 'alpha'\n"
    "\n"
    "\n"
    "class Beta:\n"
    "    def hello(self):\n"
    "        return 'beta'\n"
    "\n"
    "\n"
    "class Gamma:\n"
    "    def hello(self):\n"
    "        return 'gamma'\n"
)


def _verify_task_162(ws: Path) -> VerifyResult:
    expected = {
        "alpha.py": ("Alpha", "alpha"),
        "beta.py": ("Beta", "beta"),
        "gamma.py": ("Gamma", "gamma"),
    }
    for fname, (cls, ret) in expected.items():
        p = ws / fname
        if not p.exists():
            return VerifyResult(False, f"{fname} missing")
        text = p.read_text()
        if f"class {cls}" not in text:
            return VerifyResult(False, f"{fname} does not define class {cls}")
        for other_cls in expected.values():
            if other_cls[0] != cls and f"class {other_cls[0]}" in text:
                return VerifyResult(False, f"{fname} still contains class {other_cls[0]}")
        # quick sanity: text contains the expected return value
        if repr(ret) not in text and f"'{ret}'" not in text and f'"{ret}"' not in text:
            return VerifyResult(False, f"{fname}: return value {ret!r} not found in file")
    return VerifyResult(True, "big.py was split into alpha.py / beta.py / gamma.py")


TASK_162 = Task(
    id="task_162_split_module_by_class",
    name="Split big.py into alpha.py / beta.py / gamma.py",
    tags=("refactor", "python", "execute", "hard"),
    prompt=(
        "В файле big.py определены три класса: Alpha, Beta, Gamma — каждый с"
        " методом hello(self). Раздели файл на три отдельных модуля:\n"
        "  - alpha.py — должен содержать только определение класса Alpha;\n"
        "  - beta.py — только класс Beta;\n"
        "  - gamma.py — только класс Gamma.\n"
        "Тела методов сохрани без изменений. Сам файл big.py можно оставить"
        " или удалить — нам важны три новых файла."
    ),
    setup_files={"big.py": _BIG_MODULE_SRC},
    gold_files={
        "alpha.py": "class Alpha:\n    def hello(self):\n        return 'alpha'\n",
        "beta.py": "class Beta:\n    def hello(self):\n        return 'beta'\n",
        "gamma.py": "class Gamma:\n    def hello(self):\n        return 'gamma'\n",
    },
    verifier=_verify_task_162,
)


# 163. extract_constants
_INLINE_SRC = (
    "def retry():\n"
    "    for _ in range(5):\n"
    "        pass\n"
    "\n"
    "\n"
    "def timeout():\n"
    "    return 30\n"
)
_CONST_SRC_GOLD = (
    "from constants import MAX_RETRIES, DEFAULT_TIMEOUT\n"
    "\n"
    "\n"
    "def retry():\n"
    "    for _ in range(MAX_RETRIES):\n"
    "        pass\n"
    "\n"
    "\n"
    "def timeout():\n"
    "    return DEFAULT_TIMEOUT\n"
)
_CONSTANTS_GOLD = "MAX_RETRIES = 5\nDEFAULT_TIMEOUT = 30\n"


def _verify_task_163(ws: Path) -> VerifyResult:
    consts = ws / "constants.py"
    main = ws / "app.py"
    if not consts.exists():
        return VerifyResult(False, "constants.py missing")
    consts_text = consts.read_text()
    if "MAX_RETRIES" not in consts_text or "5" not in consts_text:
        return VerifyResult(False, f"constants.py missing MAX_RETRIES=5: {consts_text!r}")
    if "DEFAULT_TIMEOUT" not in consts_text or "30" not in consts_text:
        return VerifyResult(False, f"constants.py missing DEFAULT_TIMEOUT=30: {consts_text!r}")
    if not main.exists():
        return VerifyResult(False, "app.py missing")
    app_text = main.read_text()
    if "MAX_RETRIES" not in app_text:
        return VerifyResult(False, "app.py does not reference MAX_RETRIES")
    if "DEFAULT_TIMEOUT" not in app_text:
        return VerifyResult(False, "app.py does not reference DEFAULT_TIMEOUT")
    if "range(5)" in app_text:
        return VerifyResult(False, "app.py still contains literal range(5)")
    if "return 30" in app_text:
        return VerifyResult(False, "app.py still contains literal `return 30`")
    return VerifyResult(True, "constants moved out of app.py into constants.py")


TASK_163 = Task(
    id="task_163_extract_constants",
    name="Move magic numbers out of app.py into constants.py",
    tags=("refactor", "python", "execute", "hard"),
    prompt=(
        "В файле app.py две функции: retry() использует литерал 5 в"
        " range(5), а timeout() возвращает литерал 30. Извлеки эти константы"
        " в новый модуль constants.py:\n"
        "  - constants.py должен содержать MAX_RETRIES = 5 и"
        " DEFAULT_TIMEOUT = 30 (имена ровно такие);\n"
        "  - в app.py добавь импорт этих констант из constants и подмени"
        " 'range(5)' на 'range(MAX_RETRIES)', а 'return 30' на"
        " 'return DEFAULT_TIMEOUT'.\n"
        "Поведение функций должно остаться прежним."
    ),
    setup_files={"app.py": _INLINE_SRC},
    gold_files={"constants.py": _CONSTANTS_GOLD, "app.py": _CONST_SRC_GOLD},
    verifier=_verify_task_163,
)


# 164. add_type_hints_module
_UNTYPED_SRC = (
    "def add(a, b):\n"
    "    return a + b\n"
    "\n"
    "\n"
    "def greet(name):\n"
    "    return f'hi {name}'\n"
    "\n"
    "\n"
    "def double(x):\n"
    "    return x * 2\n"
)
_TYPED_SRC = (
    "def add(a: int, b: int) -> int:\n"
    "    return a + b\n"
    "\n"
    "\n"
    "def greet(name: str) -> str:\n"
    "    return f'hi {name}'\n"
    "\n"
    "\n"
    "def double(x: int) -> int:\n"
    "    return x * 2\n"
)
TASK_164 = Task(
    id="task_164_add_type_hints_module",
    name="Add type hints to every function in utils.py",
    tags=("refactor", "python", "execute", "hard"),
    prompt=(
        "В файле utils.py три функции: add(a, b), greet(name), double(x)."
        " Добавь к каждой аннотации типов:\n"
        "  - add: оба аргумента и возвращаемое значение — int;\n"
        "  - greet: аргумент name — str, возвращаемое значение — str;\n"
        "  - double: x — int, возвращаемое значение — int.\n"
        "Тела функций (return ...) не меняй."
    ),
    setup_files={"utils.py": _UNTYPED_SRC},
    gold_files={"utils.py": _TYPED_SRC},
    verifier=all_of(
        file_matches_regex("utils.py", r"def\s+add\(\s*a:\s*int\s*,\s*b:\s*int\s*\)\s*->\s*int\s*:"),
        file_matches_regex("utils.py", r"def\s+greet\(\s*name:\s*str\s*\)\s*->\s*str\s*:"),
        file_matches_regex("utils.py", r"def\s+double\(\s*x:\s*int\s*\)\s*->\s*int\s*:"),
        python_callable_returns("utils.py", "mod.add(2, 3)", 5),
        python_callable_returns("utils.py", "mod.greet('Bob')", "hi Bob"),
    ),
)


# 165. rename_across_files
_RENAME_FILES = {
    "a.py": "from common import old_name\n\n\ndef use_a():\n    return old_name()\n",
    "b.py": "from common import old_name\n\n\nprint(old_name())\n",
    "c.py": "import common\n\n\nprint(common.old_name())\n",
    "common.py": "def old_name():\n    return 'value'\n",
}
_RENAMED_FILES = {
    "a.py": "from common import new_name\n\n\ndef use_a():\n    return new_name()\n",
    "b.py": "from common import new_name\n\n\nprint(new_name())\n",
    "c.py": "import common\n\n\nprint(common.new_name())\n",
    "common.py": "def new_name():\n    return 'value'\n",
}
TASK_165 = Task(
    id="task_165_rename_across_files",
    name="Rename function `old_name` to `new_name` across 4 files",
    tags=("refactor", "python", "multifile", "execute", "hard"),
    prompt=(
        "В четырёх файлах — common.py, a.py, b.py, c.py — встречается функция"
        " old_name() (её определение лежит в common.py, остальные три файла"
        " её импортируют и вызывают). Переименуй её в new_name() — везде"
        " (включая определение и все импорты/вызовы). Никаких других"
        " изменений делать не надо."
    ),
    setup_files=_RENAME_FILES,
    gold_files=_RENAMED_FILES,
    verifier=all_of(
        *[file_not_contains(path, "old_name") for path in _RENAME_FILES],
        *[file_contains(path, "new_name") for path in _RENAME_FILES],
    ),
)


# 166. convert_to_dataclass
_PLAIN_CLASS_SRC = (
    "class Point:\n"
    "    def __init__(self, x, y):\n"
    "        self.x = x\n"
    "        self.y = y\n"
)
_DATACLASS_GOLD = (
    "from dataclasses import dataclass\n"
    "\n"
    "\n"
    "@dataclass\n"
    "class Point:\n"
    "    x: float\n"
    "    y: float\n"
)
TASK_166 = Task(
    id="task_166_convert_to_dataclass",
    name="Convert Point class to @dataclass",
    tags=("refactor", "python", "execute", "medium"),
    prompt=(
        "В файле point.py определён обычный класс Point с явным __init__,"
        " который принимает x и y и записывает их в self. Перепиши класс в"
        " виде dataclass: добавь импорт 'from dataclasses import dataclass',"
        " декоратор @dataclass, и объяви поля x: float и y: float (вместо"
        " ручного __init__). Поведение должно сохраниться: Point(1.0, 2.0).x"
        " == 1.0, Point(1.0, 2.0).y == 2.0."
    ),
    setup_files={"point.py": _PLAIN_CLASS_SRC},
    gold_files={"point.py": _DATACLASS_GOLD},
    verifier=all_of(
        file_contains("point.py", "from dataclasses import dataclass", "@dataclass", "class Point", "x:", "y:"),
        file_not_contains("point.py", "def __init__"),
        python_callable_returns("point.py", "mod.Point(1.0, 2.0).x", 1.0),
        python_callable_returns("point.py", "mod.Point(1.0, 2.0).y", 2.0),
    ),
)


# ---------------------------------------------------------------------------
# Group D: Algorithms + pytest (8 tasks, 167..174)
# ---------------------------------------------------------------------------

TASK_167 = Task(
    id="task_167_impl_quicksort",
    name="Implement quicksort(arr) so pytest passes",
    tags=("python", "impl", "pytest", "execute", "hard"),
    prompt=(
        "Создай файл qsort.py с функцией quicksort(arr: list[int]) -> list[int],"
        " которая возвращает отсортированный по возрастанию список целых чисел"
        " (исходный список менять не обязательно). В каталоге tests лежит"
        " test_qsort.py с проверками — тесты должны пройти."
    ),
    setup_files={
        "tests/test_qsort.py": (
            "from qsort import quicksort\n"
            "\n"
            "\n"
            "def test_empty():\n"
            "    assert quicksort([]) == []\n"
            "\n"
            "\n"
            "def test_single():\n"
            "    assert quicksort([42]) == [42]\n"
            "\n"
            "\n"
            "def test_general():\n"
            "    assert quicksort([3, 1, 4, 1, 5, 9, 2, 6, 5, 3, 5]) == [1, 1, 2, 3, 3, 4, 5, 5, 5, 6, 9]\n"
            "\n"
            "\n"
            "def test_already_sorted():\n"
            "    assert quicksort([1, 2, 3, 4]) == [1, 2, 3, 4]\n"
        ),
    },
    gold_files={
        "qsort.py": (
            "def quicksort(arr):\n"
            "    if len(arr) <= 1:\n"
            "        return list(arr)\n"
            "    pivot = arr[len(arr) // 2]\n"
            "    less = [x for x in arr if x < pivot]\n"
            "    equal = [x for x in arr if x == pivot]\n"
            "    greater = [x for x in arr if x > pivot]\n"
            "    return quicksort(less) + equal + quicksort(greater)\n"
        ),
    },
    verifier=pytest_passes("tests"),
)


TASK_168 = Task(
    id="task_168_impl_binary_search",
    name="Implement binary_search(arr, target) so pytest passes",
    tags=("python", "impl", "pytest", "execute", "hard"),
    prompt=(
        "Создай файл bsearch.py с функцией binary_search(arr: list[int], target:"
        " int) -> int, которая возвращает индекс target в отсортированном"
        " массиве arr, либо -1 если target отсутствует. В каталоге tests лежит"
        " test_bsearch.py — тесты должны пройти."
    ),
    setup_files={
        "tests/test_bsearch.py": (
            "from bsearch import binary_search\n"
            "\n"
            "\n"
            "def test_found_middle():\n"
            "    assert binary_search([1, 2, 3, 4, 5], 3) == 2\n"
            "\n"
            "\n"
            "def test_found_first():\n"
            "    assert binary_search([1, 2, 3, 4, 5], 1) == 0\n"
            "\n"
            "\n"
            "def test_found_last():\n"
            "    assert binary_search([1, 2, 3, 4, 5], 5) == 4\n"
            "\n"
            "\n"
            "def test_not_found():\n"
            "    assert binary_search([1, 2, 3, 4, 5], 42) == -1\n"
            "\n"
            "\n"
            "def test_empty():\n"
            "    assert binary_search([], 1) == -1\n"
        ),
    },
    gold_files={
        "bsearch.py": (
            "def binary_search(arr, target):\n"
            "    lo, hi = 0, len(arr) - 1\n"
            "    while lo <= hi:\n"
            "        mid = (lo + hi) // 2\n"
            "        if arr[mid] == target:\n"
            "            return mid\n"
            "        if arr[mid] < target:\n"
            "            lo = mid + 1\n"
            "        else:\n"
            "            hi = mid - 1\n"
            "    return -1\n"
        ),
    },
    verifier=pytest_passes("tests"),
)


TASK_169 = Task(
    id="task_169_impl_balanced_parens",
    name="Implement is_balanced(s) so pytest passes",
    tags=("python", "impl", "pytest", "execute", "medium"),
    prompt=(
        "Создай файл parens.py с функцией is_balanced(s: str) -> bool. Она"
        " возвращает True, если все скобки трёх видов — (), [], {} —"
        " сбалансированы (правильно открываются и закрываются в верном"
        " порядке), и False иначе. Другие символы игнорируются. В tests лежат"
        " тесты — они должны пройти."
    ),
    setup_files={
        "tests/test_parens.py": (
            "from parens import is_balanced\n"
            "\n"
            "\n"
            "def test_empty():\n"
            "    assert is_balanced('') is True\n"
            "\n"
            "\n"
            "def test_simple():\n"
            "    assert is_balanced('()') is True\n"
            "    assert is_balanced('()[]{}') is True\n"
            "\n"
            "\n"
            "def test_nested():\n"
            "    assert is_balanced('([{}])') is True\n"
            "\n"
            "\n"
            "def test_unbalanced():\n"
            "    assert is_balanced('(') is False\n"
            "    assert is_balanced('(]') is False\n"
            "    assert is_balanced('([)]') is False\n"
            "\n"
            "\n"
            "def test_with_other_chars():\n"
            "    assert is_balanced('f(x) = [1+2]') is True\n"
        ),
    },
    gold_files={
        "parens.py": (
            "def is_balanced(s):\n"
            "    pairs = {')': '(', ']': '[', '}': '{'}\n"
            "    stack = []\n"
            "    for ch in s:\n"
            "        if ch in '([{':\n"
            "            stack.append(ch)\n"
            "        elif ch in ')]}':\n"
            "            if not stack or stack[-1] != pairs[ch]:\n"
            "                return False\n"
            "            stack.pop()\n"
            "    return not stack\n"
        ),
    },
    verifier=pytest_passes("tests"),
)


TASK_170 = Task(
    id="task_170_impl_lru_cache",
    name="Implement LRUCache class so pytest passes",
    tags=("python", "impl", "pytest", "execute", "hard"),
    prompt=(
        "Создай файл lru.py с классом LRUCache(capacity). Класс должен"
        " поддерживать методы get(key) (возвращает значение или None если"
        " ключа нет) и put(key, value) (записывает пару). При превышении"
        " capacity вытесняется наименее недавно использованный элемент"
        " (Least Recently Used). И get, и put обновляют признак 'недавно"
        " использованного'. Тесты в tests/test_lru.py должны пройти."
    ),
    setup_files={
        "tests/test_lru.py": (
            "from lru import LRUCache\n"
            "\n"
            "\n"
            "def test_basic_get_put():\n"
            "    c = LRUCache(2)\n"
            "    c.put('a', 1)\n"
            "    c.put('b', 2)\n"
            "    assert c.get('a') == 1\n"
            "    assert c.get('b') == 2\n"
            "\n"
            "\n"
            "def test_miss_returns_none():\n"
            "    c = LRUCache(2)\n"
            "    assert c.get('x') is None\n"
            "\n"
            "\n"
            "def test_eviction():\n"
            "    c = LRUCache(2)\n"
            "    c.put('a', 1)\n"
            "    c.put('b', 2)\n"
            "    c.put('c', 3)\n"
            "    assert c.get('a') is None\n"
            "    assert c.get('b') == 2\n"
            "    assert c.get('c') == 3\n"
            "\n"
            "\n"
            "def test_recency_updates_on_get():\n"
            "    c = LRUCache(2)\n"
            "    c.put('a', 1)\n"
            "    c.put('b', 2)\n"
            "    c.get('a')\n"
            "    c.put('c', 3)\n"
            "    assert c.get('b') is None\n"
            "    assert c.get('a') == 1\n"
        ),
    },
    gold_files={
        "lru.py": (
            "from collections import OrderedDict\n"
            "\n"
            "\n"
            "class LRUCache:\n"
            "    def __init__(self, capacity):\n"
            "        self.capacity = capacity\n"
            "        self._data = OrderedDict()\n"
            "\n"
            "    def get(self, key):\n"
            "        if key not in self._data:\n"
            "            return None\n"
            "        self._data.move_to_end(key)\n"
            "        return self._data[key]\n"
            "\n"
            "    def put(self, key, value):\n"
            "        if key in self._data:\n"
            "            self._data.move_to_end(key)\n"
            "        self._data[key] = value\n"
            "        if len(self._data) > self.capacity:\n"
            "            self._data.popitem(last=False)\n"
        ),
    },
    verifier=pytest_passes("tests"),
)


TASK_171 = Task(
    id="task_171_impl_linked_list",
    name="Implement LinkedList class so pytest passes",
    tags=("python", "impl", "pytest", "execute", "hard"),
    prompt=(
        "Создай файл linked_list.py с классом LinkedList. Поддержи методы:\n"
        "  - append(value) — добавить элемент в конец;\n"
        "  - prepend(value) — добавить в начало;\n"
        "  - __len__() — длина списка;\n"
        "  - __iter__() — итерация в порядке от первого к последнему;\n"
        "  - to_list() — вернуть содержимое как python list.\n"
        " Тесты в tests/test_ll.py должны пройти."
    ),
    setup_files={
        "tests/test_ll.py": (
            "from linked_list import LinkedList\n"
            "\n"
            "\n"
            "def test_empty():\n"
            "    ll = LinkedList()\n"
            "    assert len(ll) == 0\n"
            "    assert ll.to_list() == []\n"
            "\n"
            "\n"
            "def test_append_order():\n"
            "    ll = LinkedList()\n"
            "    ll.append(1); ll.append(2); ll.append(3)\n"
            "    assert ll.to_list() == [1, 2, 3]\n"
            "    assert len(ll) == 3\n"
            "\n"
            "\n"
            "def test_prepend_order():\n"
            "    ll = LinkedList()\n"
            "    ll.prepend(1); ll.prepend(2); ll.prepend(3)\n"
            "    assert ll.to_list() == [3, 2, 1]\n"
            "\n"
            "\n"
            "def test_iter():\n"
            "    ll = LinkedList()\n"
            "    for v in [10, 20, 30]:\n"
            "        ll.append(v)\n"
            "    assert list(ll) == [10, 20, 30]\n"
        ),
    },
    gold_files={
        "linked_list.py": (
            "class LinkedList:\n"
            "    def __init__(self):\n"
            "        self._items = []\n"
            "\n"
            "    def append(self, value):\n"
            "        self._items.append(value)\n"
            "\n"
            "    def prepend(self, value):\n"
            "        self._items.insert(0, value)\n"
            "\n"
            "    def __len__(self):\n"
            "        return len(self._items)\n"
            "\n"
            "    def __iter__(self):\n"
            "        return iter(self._items)\n"
            "\n"
            "    def to_list(self):\n"
            "        return list(self._items)\n"
        ),
    },
    verifier=pytest_passes("tests"),
)


TASK_172 = Task(
    id="task_172_impl_tree_inorder",
    name="Implement TreeNode + inorder() so pytest passes",
    tags=("python", "impl", "pytest", "execute", "hard"),
    prompt=(
        "Создай файл tree.py с классом TreeNode(value, left=None, right=None)"
        " и свободной функцией inorder(root) -> list. Функция inorder обходит"
        " бинарное дерево в порядке: левое поддерево, корень, правое"
        " поддерево, и возвращает список значений в этом порядке. Тесты в"
        " tests/test_tree.py должны пройти."
    ),
    setup_files={
        "tests/test_tree.py": (
            "from tree import TreeNode, inorder\n"
            "\n"
            "\n"
            "def test_single():\n"
            "    assert inorder(TreeNode(1)) == [1]\n"
            "\n"
            "\n"
            "def test_left_skew():\n"
            "    root = TreeNode(3, TreeNode(2, TreeNode(1)))\n"
            "    assert inorder(root) == [1, 2, 3]\n"
            "\n"
            "\n"
            "def test_bst_inorder_sorted():\n"
            "    root = TreeNode(\n"
            "        4,\n"
            "        TreeNode(2, TreeNode(1), TreeNode(3)),\n"
            "        TreeNode(6, TreeNode(5), TreeNode(7)),\n"
            "    )\n"
            "    assert inorder(root) == [1, 2, 3, 4, 5, 6, 7]\n"
            "\n"
            "\n"
            "def test_empty():\n"
            "    assert inorder(None) == []\n"
        ),
    },
    gold_files={
        "tree.py": (
            "class TreeNode:\n"
            "    def __init__(self, value, left=None, right=None):\n"
            "        self.value = value\n"
            "        self.left = left\n"
            "        self.right = right\n"
            "\n"
            "\n"
            "def inorder(root):\n"
            "    if root is None:\n"
            "        return []\n"
            "    return inorder(root.left) + [root.value] + inorder(root.right)\n"
        ),
    },
    verifier=pytest_passes("tests"),
)


TASK_173 = Task(
    id="task_173_impl_anagram",
    name="Implement is_anagram(a, b) so pytest passes",
    tags=("python", "impl", "pytest", "execute", "medium"),
    prompt=(
        "Создай файл anagram.py с функцией is_anagram(a: str, b: str) -> bool."
        " True, если a — анаграмма b (те же буквы в любом порядке, регистр"
        " неважен, пробелы игнорируются). False иначе. Тесты в"
        " tests/test_anagram.py должны пройти."
    ),
    setup_files={
        "tests/test_anagram.py": (
            "from anagram import is_anagram\n"
            "\n"
            "\n"
            "def test_yes():\n"
            "    assert is_anagram('listen', 'silent') is True\n"
            "    assert is_anagram('Astronomer', 'Moon starer') is True\n"
            "\n"
            "\n"
            "def test_no():\n"
            "    assert is_anagram('hello', 'world') is False\n"
            "    assert is_anagram('abc', 'ab') is False\n"
            "\n"
            "\n"
            "def test_empty():\n"
            "    assert is_anagram('', '') is True\n"
        ),
    },
    gold_files={
        "anagram.py": (
            "def is_anagram(a, b):\n"
            "    norm = lambda s: sorted(s.lower().replace(' ', ''))\n"
            "    return norm(a) == norm(b)\n"
        ),
    },
    verifier=pytest_passes("tests"),
)


TASK_174 = Task(
    id="task_174_impl_two_sum",
    name="Implement two_sum(nums, target) so pytest passes",
    tags=("python", "impl", "pytest", "execute", "medium"),
    prompt=(
        "Создай файл twosum.py с функцией two_sum(nums: list[int], target:"
        " int) -> tuple[int, int] | None. Если в списке есть два индекса i, j"
        " (i < j), при которых nums[i] + nums[j] == target — вернуть кортеж"
        " (i, j). Если такой пары нет — вернуть None. Тесты в"
        " tests/test_twosum.py должны пройти."
    ),
    setup_files={
        "tests/test_twosum.py": (
            "from twosum import two_sum\n"
            "\n"
            "\n"
            "def test_simple():\n"
            "    assert two_sum([2, 7, 11, 15], 9) == (0, 1)\n"
            "\n"
            "\n"
            "def test_later():\n"
            "    assert two_sum([3, 2, 4], 6) == (1, 2)\n"
            "\n"
            "\n"
            "def test_no_pair():\n"
            "    assert two_sum([1, 2, 3], 100) is None\n"
            "\n"
            "\n"
            "def test_empty():\n"
            "    assert two_sum([], 0) is None\n"
        ),
    },
    gold_files={
        "twosum.py": (
            "def two_sum(nums, target):\n"
            "    seen = {}\n"
            "    for j, v in enumerate(nums):\n"
            "        need = target - v\n"
            "        if need in seen:\n"
            "            return (seen[need], j)\n"
            "        seen[v] = j\n"
            "    return None\n"
        ),
    },
    verifier=pytest_passes("tests"),
)


# ---------------------------------------------------------------------------
# Group E: Data analysis (8 tasks, 175..182)
# ---------------------------------------------------------------------------

_NUMS_LIST = [12, 18, 25, 31, 8, 22, 27, 19, 33, 16, 21, 24, 30, 14, 26, 29, 11, 20, 28, 23]
_NUMS_CSV = "value\n" + "\n".join(str(n) for n in _NUMS_LIST) + "\n"


# 175. csv_stats_basic
_STATS_GOLD = {
    "mean": round(sum(_NUMS_LIST) / len(_NUMS_LIST), 2),
    "median": statistics.median(_NUMS_LIST),
    "min": min(_NUMS_LIST),
    "max": max(_NUMS_LIST),
}


def _verify_task_175(ws: Path) -> VerifyResult:
    p = ws / "stats.json"
    if not p.exists():
        return VerifyResult(False, "stats.json missing")
    try:
        data = json.loads(p.read_text())
    except json.JSONDecodeError as exc:
        return VerifyResult(False, f"stats.json invalid JSON: {exc}")
    for key, want in _STATS_GOLD.items():
        got = data.get(key)
        if got is None:
            return VerifyResult(False, f"stats.json missing key {key!r}")
        try:
            if abs(float(got) - float(want)) > 0.01:
                return VerifyResult(False, f"stats.json[{key}] = {got!r}, expected {want!r}")
        except (TypeError, ValueError):
            return VerifyResult(False, f"stats.json[{key}] is not numeric: {got!r}")
    return VerifyResult(True, "stats.json has mean, median, min, max within tolerance")


TASK_175 = Task(
    id="task_175_csv_stats_basic",
    name="Compute mean/median/min/max of a CSV column",
    tags=("csv", "stats", "compute", "execute", "hard"),
    prompt=(
        "В файле numbers.csv 20 строк данных (одна колонка value, плюс"
        " заголовок). Посчитай по этой колонке четыре значения: mean (среднее"
        " арифметическое), median (медиана как (a+b)/2 для чётного количества"
        " элементов), min, max. Сохрани результат в stats.json как объект с"
        " ключами 'mean', 'median', 'min', 'max'. mean можно округлить до 2"
        " знаков после запятой; median не округлять (может быть дробным,"
        " например 22.5); min и max — целые числа."
    ),
    setup_files={"numbers.csv": _NUMS_CSV},
    gold_files={"stats.json": json.dumps(_STATS_GOLD) + "\n"},
    verifier=_verify_task_175,
)


# 176. csv_rolling_avg
_ROLLING_INPUT = [10, 20, 30, 40, 50, 60, 70, 80]
_ROLLING_CSV = "value\n" + "\n".join(str(n) for n in _ROLLING_INPUT) + "\n"
_ROLLING_GOLD = []  # list of (value, rolling_avg or "")
for i, v in enumerate(_ROLLING_INPUT):
    if i < 2:
        _ROLLING_GOLD.append((v, ""))
    else:
        win = _ROLLING_INPUT[i - 2 : i + 1]
        _ROLLING_GOLD.append((v, round(sum(win) / 3, 2)))
_ROLLING_OUT = "value,rolling_avg\n" + "".join(
    f"{v},{ra}\n" for v, ra in _ROLLING_GOLD
)


def _verify_task_176(ws: Path) -> VerifyResult:
    p = ws / "rolling.csv"
    if not p.exists():
        return VerifyResult(False, "rolling.csv missing")
    lines = [line for line in p.read_text().splitlines() if line.strip()]
    if not lines or lines[0].strip().lower() not in {"value,rolling_avg", "value, rolling_avg"}:
        return VerifyResult(False, f"rolling.csv header is {lines[0]!r}, expected 'value,rolling_avg'")
    data_lines = lines[1:]
    if len(data_lines) != len(_ROLLING_INPUT):
        return VerifyResult(False, f"rolling.csv has {len(data_lines)} data rows, expected {len(_ROLLING_INPUT)}")
    for i, line in enumerate(data_lines):
        parts = [x.strip() for x in line.split(",")]
        try:
            val = int(parts[0])
        except ValueError:
            return VerifyResult(False, f"rolling.csv row {i}: value is not int: {parts[0]!r}")
        if val != _ROLLING_INPUT[i]:
            return VerifyResult(False, f"rolling.csv row {i}: value {val} != expected {_ROLLING_INPUT[i]}")
        expected_avg = _ROLLING_GOLD[i][1]
        if expected_avg == "":
            if parts[1] not in ("", "NaN", "nan", "null"):
                return VerifyResult(False, f"rolling.csv row {i}: expected empty rolling_avg, got {parts[1]!r}")
        else:
            try:
                got = float(parts[1])
            except ValueError:
                return VerifyResult(False, f"rolling.csv row {i}: rolling_avg not numeric: {parts[1]!r}")
            if abs(got - float(expected_avg)) > 0.01:
                return VerifyResult(False, f"rolling.csv row {i}: rolling_avg {got} != {expected_avg}")
    return VerifyResult(True, "rolling.csv has correct rolling averages with empty first two cells")


TASK_176 = Task(
    id="task_176_csv_rolling_avg",
    name="Add a 3-window rolling average column to numbers.csv",
    tags=("csv", "stats", "execute", "hard"),
    prompt=(
        "В файле numbers.csv 8 строк данных в колонке value. Сделай файл"
        " rolling.csv с двумя колонками: value и rolling_avg. rolling_avg —"
        " это среднее значений value в скользящем окне из трёх последних"
        " строк (текущей и двух предыдущих). Для первых двух строк, у"
        " которых ещё нет трёх значений, оставь поле rolling_avg пустым (то"
        " есть строка имеет вид 'value,'). Значение rolling_avg округляй до"
        " 2 знаков после запятой."
    ),
    setup_files={"numbers.csv": _ROLLING_CSV},
    gold_files={"rolling.csv": _ROLLING_OUT},
    verifier=_verify_task_176,
)


# 177. csv_histogram_json
_HIST_VALUES = [3, 7, 14, 22, 28, 35, 41, 48, 55, 62, 71, 79, 88, 95, 17, 33, 50, 66, 82, 99]
_HIST_BINS = [0, 20, 40, 60, 80, 100]
_HIST_LABELS = ["0-20", "20-40", "40-60", "60-80", "80-100"]


def _bin_for(v: int) -> str:
    for i in range(len(_HIST_BINS) - 1):
        lo, hi = _HIST_BINS[i], _HIST_BINS[i + 1]
        if lo <= v < hi or (i == len(_HIST_BINS) - 2 and v == hi):
            return _HIST_LABELS[i]
    raise ValueError(v)


_HIST_GOLD: dict[str, int] = {label: 0 for label in _HIST_LABELS}
for v in _HIST_VALUES:
    _HIST_GOLD[_bin_for(v)] += 1
_HIST_CSV = "value\n" + "\n".join(str(n) for n in _HIST_VALUES) + "\n"
TASK_177 = Task(
    id="task_177_csv_histogram",
    name="Bucket values into a 5-bin histogram JSON",
    tags=("csv", "stats", "json", "execute", "hard"),
    prompt=(
        "В файле values.csv 20 строк данных в колонке value (числа от 0 до"
        " 99). Разнеси значения по пяти бинам и сохрани результат в"
        " histogram.json как объект, где ключи — метки бинов ('0-20', '20-40',"
        " '40-60', '60-80', '80-100'), а значения — количество элементов в"
        " каждом бине. Включи в бин '0-20' значения 0 <= v < 20, в '20-40'"
        " 20 <= v < 40, и так далее; верхнюю границу '80-100' трактуй"
        " включительно (80 <= v <= 100)."
    ),
    setup_files={"values.csv": _HIST_CSV},
    gold_files={"histogram.json": json.dumps(_HIST_GOLD) + "\n"},
    verifier=_json_file_matches_loose("histogram.json", _HIST_GOLD),
)


# 178. csv_pivot_count
_PIVOT_ROWS = [
    ("moscow", "food"), ("moscow", "tech"), ("berlin", "food"),
    ("moscow", "food"), ("berlin", "tech"), ("paris", "food"),
    ("paris", "food"), ("moscow", "tech"), ("berlin", "food"),
    ("paris", "tech"), ("moscow", "food"), ("berlin", "food"),
]
_PIVOT_INPUT = "city,category\n" + "".join(f"{c},{k}\n" for c, k in _PIVOT_ROWS)
_PIVOT_COUNT: dict[tuple[str, str], int] = Counter(_PIVOT_ROWS)
# pivot.csv: rows = sorted cities, cols = sorted categories
_pivot_cities = sorted({c for c, _ in _PIVOT_ROWS})
_pivot_cats = sorted({k for _, k in _PIVOT_ROWS})
_PIVOT_OUT_HEADER = "city," + ",".join(_pivot_cats) + "\n"
_PIVOT_OUT_BODY = "".join(
    city + "," + ",".join(str(_PIVOT_COUNT.get((city, cat), 0)) for cat in _pivot_cats) + "\n"
    for city in _pivot_cities
)
_PIVOT_OUT = _PIVOT_OUT_HEADER + _PIVOT_OUT_BODY
TASK_178 = Task(
    id="task_178_csv_pivot_count",
    name="Pivot transactions.csv into pivot.csv (rows=city, cols=category)",
    tags=("csv", "groupby", "execute", "hard"),
    prompt=(
        "В файле transactions.csv 12 строк данных (city,category). Сформируй"
        " сводную таблицу pivot.csv:\n"
        "  - заголовок: 'city,<cat1>,<cat2>,...' (категории отсортированы по"
        " алфавиту);\n"
        "  - каждая строка данных — название города (в алфавитном порядке)"
        " и через запятую количество транзакций по каждой категории;\n"
        "  - если для пары (city, category) нет строк — пиши 0."
    ),
    setup_files={"transactions.csv": _PIVOT_INPUT},
    gold_files={"pivot.csv": _PIVOT_OUT},
    verifier=file_text_equals("pivot.csv", _PIVOT_OUT),
)


# 179. csv_zscore_outliers
_Z_VALUES = [10, 12, 11, 9, 10, 100, 8, 11, 12, 9, 10, 11, 13, 8, 10]
_Z_MEAN = sum(_Z_VALUES) / len(_Z_VALUES)
_Z_STD = math.sqrt(sum((v - _Z_MEAN) ** 2 for v in _Z_VALUES) / len(_Z_VALUES))
_Z_OUTLIERS = [v for v in _Z_VALUES if abs((v - _Z_MEAN) / _Z_STD) > 2]


def _verify_task_179(ws: Path) -> VerifyResult:
    p = ws / "outliers.csv"
    if not p.exists():
        return VerifyResult(False, "outliers.csv missing")
    lines = [line.strip() for line in p.read_text().splitlines() if line.strip()]
    if not lines:
        return VerifyResult(False, "outliers.csv empty")
    if lines[0].lower() not in {"value"}:
        return VerifyResult(False, f"outliers.csv first line is {lines[0]!r}, expected 'value'")
    try:
        actual = sorted(int(line) for line in lines[1:])
    except ValueError:
        return VerifyResult(False, f"outliers.csv has non-int rows: {lines[1:]!r}")
    expected = sorted(_Z_OUTLIERS)
    if actual == expected:
        return VerifyResult(True, f"outliers.csv has the expected outliers {expected}")
    return VerifyResult(False, f"outliers.csv {actual!r} != expected {expected!r}")


TASK_179 = Task(
    id="task_179_csv_zscore_outliers",
    name="Filter z-score outliers (|z|>2) into outliers.csv",
    tags=("csv", "stats", "execute", "hard"),
    prompt=(
        "В файле values.csv 15 строк данных (одна колонка value). Сохрани в"
        " файл outliers.csv (с тем же заголовком 'value') только те значения,"
        " у которых модуль z-score больше 2 — то есть |v - mean| / std > 2,"
        " где mean и std считаются по всем 15 значениям. std считай"
        " как корень из дисперсии популяции (делитель N, а не N-1)."
    ),
    setup_files={"values.csv": "value\n" + "\n".join(str(n) for n in _Z_VALUES) + "\n"},
    gold_files={"outliers.csv": "value\n" + "\n".join(str(v) for v in _Z_OUTLIERS) + "\n"},
    verifier=_verify_task_179,
)


# 180. csv_percentiles
_PERC_VALUES = list(range(1, 101))  # 1..100; percentiles trivial


def _percentile(sorted_values: list[int], p: float) -> float:
    """Linear interpolation percentile (same convention as numpy default)."""
    if not sorted_values:
        raise ValueError("empty")
    if p <= 0:
        return float(sorted_values[0])
    if p >= 100:
        return float(sorted_values[-1])
    rank = (p / 100) * (len(sorted_values) - 1)
    lo = int(rank)
    frac = rank - lo
    if lo + 1 >= len(sorted_values):
        return float(sorted_values[-1])
    return sorted_values[lo] + frac * (sorted_values[lo + 1] - sorted_values[lo])


_PERC_GOLD = {
    f"p{p}": round(_percentile(sorted(_PERC_VALUES), p), 2)
    for p in (25, 50, 75, 90, 95)
}


def _verify_task_180(ws: Path) -> VerifyResult:
    p = ws / "percentiles.json"
    if not p.exists():
        return VerifyResult(False, "percentiles.json missing")
    try:
        data = json.loads(p.read_text())
    except json.JSONDecodeError as exc:
        return VerifyResult(False, f"percentiles.json invalid JSON: {exc}")
    for key, want in _PERC_GOLD.items():
        got = data.get(key)
        if got is None:
            return VerifyResult(False, f"percentiles.json missing key {key!r}")
        try:
            if abs(float(got) - float(want)) > 0.5:
                return VerifyResult(False, f"percentiles.json[{key}] = {got!r}, expected {want!r}")
        except (TypeError, ValueError):
            return VerifyResult(False, f"percentiles.json[{key}] not numeric: {got!r}")
    return VerifyResult(True, "percentiles.json values within tolerance of expected")


TASK_180 = Task(
    id="task_180_csv_percentiles",
    name="Compute p25/p50/p75/p90/p95 of a CSV column",
    tags=("csv", "stats", "json", "execute", "hard"),
    prompt=(
        "В файле values.csv 100 строк данных (целые числа от 1 до 100 в"
        " произвольном порядке) в колонке value. Посчитай для них пять"
        " перцентилей: 25-й, 50-й, 75-й, 90-й, 95-й. Сохрани результат в"
        " percentiles.json как объект с ключами 'p25', 'p50', 'p75', 'p90',"
        " 'p95' и числовыми значениями (можно округлять до целых или с"
        " точностью до 0.5). Допустимо использовать numpy/statistics."
    ),
    setup_files={"values.csv": "value\n" + "\n".join(str(v) for v in sorted(_PERC_VALUES, reverse=True)) + "\n"},
    gold_files={"percentiles.json": json.dumps(_PERC_GOLD) + "\n"},
    verifier=_verify_task_180,
)


# 181. csv_cumsum
_CUMSUM_INPUT = [10, 20, 30, 40, 50]
_CUMSUM_INPUT_CSV = "value\n" + "\n".join(str(v) for v in _CUMSUM_INPUT) + "\n"
_CUMSUM_GOLD_LINES = ["value,cumsum"]
_total = 0
for v in _CUMSUM_INPUT:
    _total += v
    _CUMSUM_GOLD_LINES.append(f"{v},{_total}")
_CUMSUM_GOLD = "\n".join(_CUMSUM_GOLD_LINES) + "\n"
TASK_181 = Task(
    id="task_181_csv_cumsum",
    name="Add a cumulative-sum column to numbers.csv",
    tags=("csv", "stats", "execute", "medium"),
    prompt=(
        "В файле numbers.csv 5 строк данных в колонке value. Сделай файл"
        " cumulative.csv с двумя колонками 'value,cumsum', в котором cumsum"
        " каждой строки — это сумма value этой и всех предыдущих строк."
    ),
    setup_files={"numbers.csv": _CUMSUM_INPUT_CSV},
    gold_files={"cumulative.csv": _CUMSUM_GOLD},
    verifier=file_text_equals("cumulative.csv", _CUMSUM_GOLD),
)


# 182. csv_group_agg_multi
_GA_ROWS = [
    ("food", 100), ("food", 200), ("food", 300),
    ("tech", 50), ("tech", 150),
    ("books", 80), ("books", 120), ("books", 100), ("books", 60),
]
_GA_INPUT = "category,amount\n" + "".join(f"{c},{a}\n" for c, a in _GA_ROWS)
_GA_GROUPS: dict[str, list[int]] = defaultdict(list)
for c, a in _GA_ROWS:
    _GA_GROUPS[c].append(a)
_GA_GOLD = {
    cat: {"sum": sum(vs), "mean": round(sum(vs) / len(vs), 2), "count": len(vs)}
    for cat, vs in _GA_GROUPS.items()
}


def _verify_task_182(ws: Path) -> VerifyResult:
    p = ws / "agg.json"
    if not p.exists():
        return VerifyResult(False, "agg.json missing")
    try:
        data = json.loads(p.read_text())
    except json.JSONDecodeError as exc:
        return VerifyResult(False, f"agg.json invalid JSON: {exc}")
    for cat, want in _GA_GOLD.items():
        got = data.get(cat)
        if not isinstance(got, dict):
            return VerifyResult(False, f"agg.json[{cat}] is {got!r}, expected an object")
        for key in ("sum", "mean", "count"):
            try:
                if abs(float(got.get(key)) - float(want[key])) > 0.01:
                    return VerifyResult(False, f"agg.json[{cat}][{key}] = {got.get(key)!r}, expected {want[key]!r}")
            except (TypeError, ValueError):
                return VerifyResult(False, f"agg.json[{cat}][{key}] not numeric: {got.get(key)!r}")
    return VerifyResult(True, "agg.json has sum/mean/count per category within tolerance")


TASK_182 = Task(
    id="task_182_csv_group_agg",
    name="Aggregate sales.csv: per-category sum/mean/count",
    tags=("csv", "groupby", "json", "execute", "hard"),
    prompt=(
        "В файле sales.csv 9 строк данных (category,amount). Сформируй файл"
        " agg.json — объект, в котором ключи — категории, а значения —"
        " вложенные объекты с тремя полями: sum (сумма amount по категории),"
        " mean (среднее amount, округление до 2 знаков), count (количество"
        " строк)."
    ),
    setup_files={"sales.csv": _GA_INPUT},
    gold_files={"agg.json": json.dumps(_GA_GOLD) + "\n"},
    verifier=_verify_task_182,
)


# ---------------------------------------------------------------------------
# Group F: Multi-format / cross-file (5 tasks, 183..187)
# ---------------------------------------------------------------------------

# 183. xml_to_json
_XML_INPUT = (
    "<?xml version='1.0' encoding='UTF-8'?>\n"
    "<users>\n"
    "  <user id='1'><name>Alice</name><age>30</age></user>\n"
    "  <user id='2'><name>Bob</name><age>25</age></user>\n"
    "  <user id='3'><name>Carol</name><age>40</age></user>\n"
    "</users>\n"
)
_XML_GOLD = [
    {"id": 1, "name": "Alice", "age": 30},
    {"id": 2, "name": "Bob", "age": 25},
    {"id": 3, "name": "Carol", "age": 40},
]


def _verify_task_183(ws: Path) -> VerifyResult:
    p = ws / "users.json"
    if not p.exists():
        return VerifyResult(False, "users.json missing")
    try:
        data = json.loads(p.read_text())
    except json.JSONDecodeError as exc:
        return VerifyResult(False, f"users.json invalid JSON: {exc}")
    if not isinstance(data, list) or len(data) != len(_XML_GOLD):
        return VerifyResult(False, f"users.json: expected list of 3, got {data!r}")
    expected_by_id = {row["id"]: row for row in _XML_GOLD}
    for row in data:
        rid_raw = row.get("id")
        try:
            rid = int(rid_raw)
        except (TypeError, ValueError):
            return VerifyResult(False, f"users.json row id not int-like: {rid_raw!r}")
        want = expected_by_id.get(rid)
        if want is None:
            return VerifyResult(False, f"users.json has unexpected id {rid!r}")
        if row.get("name") != want["name"]:
            return VerifyResult(False, f"users.json id={rid}: name {row.get('name')!r} != {want['name']!r}")
        try:
            if int(row.get("age")) != want["age"]:
                return VerifyResult(False, f"users.json id={rid}: age {row.get('age')!r} != {want['age']!r}")
        except (TypeError, ValueError):
            return VerifyResult(False, f"users.json id={rid}: age not int-like: {row.get('age')!r}")
    return VerifyResult(True, "users.json has the 3 users with correct fields")


TASK_183 = Task(
    id="task_183_xml_to_json",
    name="Convert users.xml to users.json",
    tags=("xml", "json", "convert", "execute", "hard"),
    prompt=(
        "В файле users.xml корневой элемент <users>, у каждого <user> есть"
        " атрибут id и вложенные элементы <name> и <age>. Преобразуй файл в"
        " users.json как JSON-массив объектов с полями id (число), name"
        " (строка), age (число) — по одному объекту на пользователя. Порядок"
        " не важен. Используй xml.etree.ElementTree или любой другой подход."
    ),
    setup_files={"users.xml": _XML_INPUT},
    gold_files={"users.json": json.dumps(_XML_GOLD) + "\n"},
    verifier=_verify_task_183,
)


# 184. md_toc_gen
_MD_INPUT = (
    "# Intro\n\nText\n\n"
    "## Installation\n\nText\n\n"
    "## Usage\n\nText\n\n"
    "### Basic\n\nText\n\n"
    "### Advanced\n\nText\n\n"
    "## License\n\nText\n"
)
_TOC_GOLD = (
    "- Intro\n"
    "  - Installation\n"
    "  - Usage\n"
    "    - Basic\n"
    "    - Advanced\n"
    "  - License\n"
)


def _verify_task_184(ws: Path) -> VerifyResult:
    p = ws / "toc.md"
    if not p.exists():
        return VerifyResult(False, "toc.md missing")
    actual = p.read_text().strip()
    expected = _TOC_GOLD.strip()
    if actual == expected:
        return VerifyResult(True, "toc.md matches expected TOC")
    return VerifyResult(False, f"toc.md content differs\n--- got ---\n{actual}\n--- exp ---\n{expected}")


TASK_184 = Task(
    id="task_184_md_toc_gen",
    name="Generate table of contents from doc.md headers",
    tags=("markdown", "compute", "execute", "hard"),
    prompt=(
        "В файле doc.md есть заголовки уровней 1, 2 и 3 (строки, начинающиеся"
        " с одной, двух или трёх решёток соответственно, после которых идёт"
        " пробел и текст заголовка). Сделай файл toc.md, в котором содержание"
        " представлено вложенным маркдаун-списком: для каждого заголовка"
        " отступ в 2 пробела на уровень вложенности, маркер '- ' и текст"
        " заголовка. Уровень 1 — без отступа, уровень 2 — 2 пробела, уровень"
        " 3 — 4 пробела. Порядок заголовков сохрани из исходного файла."
    ),
    setup_files={"doc.md": _MD_INPUT},
    gold_files={"toc.md": _TOC_GOLD},
    verifier=_verify_task_184,
)


# 185. md_frontmatter
_POST_MD = (
    "---\n"
    "title: Hello World\n"
    "author: Alice\n"
    "tags: [intro, hello]\n"
    "---\n"
    "\n"
    "Post body here.\n"
)
_FRONT_GOLD = {"title": "Hello World", "author": "Alice", "tags": ["intro", "hello"]}


def _verify_task_185(ws: Path) -> VerifyResult:
    p = ws / "frontmatter.json"
    if not p.exists():
        return VerifyResult(False, "frontmatter.json missing")
    try:
        data = json.loads(p.read_text())
    except json.JSONDecodeError as exc:
        return VerifyResult(False, f"frontmatter.json invalid JSON: {exc}")
    if data == _FRONT_GOLD:
        return VerifyResult(True, "frontmatter.json matches expected mapping")
    return VerifyResult(False, f"frontmatter.json {data!r} differs from {_FRONT_GOLD!r}")


TASK_185 = Task(
    id="task_185_md_frontmatter",
    name="Parse YAML front-matter from post.md into frontmatter.json",
    tags=("markdown", "yaml", "json", "execute", "hard"),
    prompt=(
        "Файл post.md начинается с YAML-фронтматтера: блок между двумя"
        " строками '---'. Внутри блока — валидный YAML с полями title (строка),"
        " author (строка) и tags (список строк). Сохрани этот фронтматтер как"
        " JSON-объект в файле frontmatter.json. Тело поста (после второго"
        " '---') игнорируется."
    ),
    setup_files={"post.md": _POST_MD},
    gold_files={"frontmatter.json": json.dumps(_FRONT_GOLD) + "\n"},
    verifier=_verify_task_185,
)


# 186. find_func_usages
_USAGE_FILES = {
    "src/lib.py": "def target():\n    return 42\n",
    "src/a.py": "from lib import target\n\n\nprint(target())\n",
    "src/b.py": "from lib import target\n\n\nx = target()\nprint(x)\n",
    "src/c.py": "# does not call target\nprint('hi')\n",
    "src/d.py": "from lib import target as t\n\n\nprint(t())\n",
}
# `target` substring appears in: a.py (2x: import + call), b.py (2x), d.py (1x — only import, since renamed).
# But the question is about call sites of `target()`. Let me re-check carefully:
# a.py calls target() once.
# b.py calls target() once.
# c.py — no calls.
# d.py — uses alias `t()`, not `target()` — so 0 calls.
# Total call sites of `target()` (as substring `target(`) in non-defining files: 2.
_USAGE_CALL_FILES = sorted({"a.py", "b.py"})


def _verify_task_186(ws: Path) -> VerifyResult:
    p = ws / "callers.txt"
    if not p.exists():
        return VerifyResult(False, "callers.txt missing")
    raw = [line.strip() for line in p.read_text().splitlines() if line.strip()]
    normalised = sorted({line.split("/")[-1] for line in raw})
    if normalised == _USAGE_CALL_FILES:
        return VerifyResult(True, "callers.txt has the two files that call target()")
    return VerifyResult(False, f"callers.txt {normalised!r} differs from {_USAGE_CALL_FILES!r}")


TASK_186 = Task(
    id="task_186_find_call_sites",
    name="List files under src/ that call target()",
    tags=("grep", "search", "execute", "hard"),
    prompt=(
        "В каталоге src лежит несколько .py-файлов. В lib.py определена"
        " функция target(). Найди все файлы (кроме самого lib.py), в которых"
        " встречается её непосредственный вызов 'target(' (с открывающей"
        " круглой скобкой). Запиши имена этих файлов в callers.txt по одному"
        " на строку. Имя файла без префикса каталога допустимо, как и"
        " относительный путь с 'src/'. Порядок не важен."
    ),
    setup_files=_USAGE_FILES,
    gold_files={"callers.txt": "\n".join(_USAGE_CALL_FILES) + "\n"},
    verifier=_verify_task_186,
)


# 187. detect_dead_funcs
_DEAD_FILES = {
    "code/utils.py": (
        "def used_once():\n    return 1\n\n\n"
        "def used_twice():\n    return 2\n\n\n"
        "def never_called():\n    return 3\n\n\n"
        "def also_dead():\n    return 4\n"
    ),
    "code/main.py": (
        "from utils import used_once, used_twice\n\n\n"
        "print(used_once())\n"
        "print(used_twice())\n"
        "print(used_twice())\n"
    ),
}
_DEAD_FUNCS = {"never_called", "also_dead"}


def _verify_task_187(ws: Path) -> VerifyResult:
    p = next(
        (candidate for candidate in (ws / "dead.txt", ws / "code" / "dead.txt") if candidate.exists()),
        None,
    )
    if p is None:
        return VerifyResult(False, "dead.txt missing")
    raw = {line.strip() for line in p.read_text().splitlines() if line.strip()}
    if raw == _DEAD_FUNCS:
        return VerifyResult(True, "dead.txt lists the two unused functions")
    return VerifyResult(False, f"dead.txt {sorted(raw)} differs from {sorted(_DEAD_FUNCS)}")


TASK_187 = Task(
    id="task_187_dead_functions",
    name="Find functions defined in utils.py but never called from main.py",
    tags=("grep", "search", "execute", "hard"),
    prompt=(
        "В каталоге code два файла:\n"
        "  - utils.py — определяет четыре функции (used_once, used_twice,"
        " never_called, also_dead);\n"
        "  - main.py — импортирует и вызывает что-то из них.\n"
        "Сделай файл dead.txt со списком имён функций, которые ОПРЕДЕЛЕНЫ в"
        " utils.py, но НИ РАЗУ не вызываются из main.py (то есть в main.py не"
        " встречается подстрока '<имя>('). Записывай по одному имени на"
        " строку, порядок не важен."
    ),
    setup_files=_DEAD_FILES,
    gold_files={"dead.txt": "\n".join(sorted(_DEAD_FUNCS)) + "\n"},
    verifier=_verify_task_187,
)


# ---------------------------------------------------------------------------
# Group G: Hard composite (8 tasks, 188..195)
# ---------------------------------------------------------------------------

# 188. csv_join_3way
_J3_USERS = "user_id,name\n1,Alice\n2,Bob\n3,Carol\n"
_J3_ORDERS = "order_id,user_id,product_id\n10,1,p1\n11,1,p2\n12,2,p1\n13,3,p3\n"
_J3_PRODUCTS = "product_id,price\np1,100\np2,50\np3,200\n"
# Inner join in order of orders.csv:
_J3_GOLD = (
    "order_id,name,product_id,price\n"
    "10,Alice,p1,100\n"
    "11,Alice,p2,50\n"
    "12,Bob,p1,100\n"
    "13,Carol,p3,200\n"
)
TASK_188 = Task(
    id="task_188_csv_three_way_join",
    name="Three-way join: orders + users + products",
    tags=("csv", "join", "pipeline", "execute", "hard"),
    prompt=(
        "Три CSV-файла лежат в корне рабочей директории:\n"
        "  - users.csv: user_id,name\n"
        "  - orders.csv: order_id,user_id,product_id\n"
        "  - products.csv: product_id,price\n"
        "Сделай файл joined.csv с заголовком 'order_id,name,product_id,price'"
        " — для каждой строки orders.csv (в исходном порядке) добавь имя"
        " пользователя из users.csv и цену продукта из products.csv."
    ),
    setup_files={"users.csv": _J3_USERS, "orders.csv": _J3_ORDERS, "products.csv": _J3_PRODUCTS},
    gold_files={"joined.csv": _J3_GOLD},
    verifier=file_text_equals("joined.csv", _J3_GOLD),
)


# 189. logs_aggregate_hourly
def _make_access_log_189() -> tuple[str, dict[str, int]]:
    hours = {"08": 5, "09": 12, "10": 20, "11": 8, "12": 3}
    lines = []
    for h, count in hours.items():
        for _ in range(count):
            lines.append(
                f'10.0.0.1 - - [13/May/2026:{h}:00:00 +0000] '
                f'"GET /api HTTP/1.1" 200 100'
            )
    return "\n".join(lines) + "\n", hours


_HOURLY_LOG, _HOURLY_COUNTS = _make_access_log_189()
_HOURLY_CSV = "hour,count\n" + "".join(f"{h},{c}\n" for h, c in sorted(_HOURLY_COUNTS.items()))


def _verify_task_189(ws: Path) -> VerifyResult:
    p = ws / "hourly.csv"
    if not p.exists():
        return VerifyResult(False, "hourly.csv missing")
    import csv as _csv  # noqa: PLC0415

    rows = list(_csv.DictReader(io.StringIO(p.read_text())))
    if not rows:
        return VerifyResult(False, "hourly.csv has no data")
    actual = {r["hour"]: int(r["count"]) for r in rows}
    if actual == _HOURLY_COUNTS:
        return VerifyResult(True, "hourly.csv counts match")
    return VerifyResult(False, f"hourly.csv {actual!r} differs from {_HOURLY_COUNTS!r}")


TASK_189 = Task(
    id="task_189_log_hourly_counts",
    name="Group access.log requests by hour, write hourly.csv",
    tags=("logs", "csv", "groupby", "execute", "hard"),
    prompt=(
        "В файле access.log около 48 строк в формате Apache combined log."
        " В каждой строке временная метка вида [DD/Mon/YYYY:HH:MM:SS +ZZZZ]."
        " Сгруппируй запросы по часу (две цифры HH) и сохрани результат в"
        " hourly.csv с заголовком 'hour,count' — по строке на каждый"
        " встретившийся час. Порядок строк не важен."
    ),
    setup_files={"access.log": _HOURLY_LOG},
    gold_files={"hourly.csv": _HOURLY_CSV},
    verifier=_verify_task_189,
)


# 190. multi_csv_concat_dedupe_sort
_CSVS_190 = {
    "data/part_a.csv": "id,name\n2,Bob\n1,Alice\n3,Carol\n",
    "data/part_b.csv": "id,name\n4,Dave\n1,Alice\n5,Eve\n",
    "data/part_c.csv": "id,name\n3,Carol\n6,Frank\n",
}
# Concat, dedupe by full row, sort by id asc:
_CONCAT_ROWS = sorted(
    {(1, "Alice"), (2, "Bob"), (3, "Carol"), (4, "Dave"), (5, "Eve"), (6, "Frank")}
)
_CONCAT_GOLD = "id,name\n" + "".join(f"{i},{n}\n" for i, n in _CONCAT_ROWS)


def _verify_task_190(ws: Path) -> VerifyResult:
    p = next(
        (candidate for candidate in (ws / "merged.csv", ws / "data" / "merged.csv") if candidate.exists()),
        None,
    )
    if p is None:
        return VerifyResult(False, "merged.csv missing")
    actual = p.read_text(encoding="utf-8").replace("\r\n", "\n").strip()
    expected = _CONCAT_GOLD.strip()
    if actual == expected:
        return VerifyResult(True, "merged.csv matches expected content")
    return VerifyResult(False, f"merged.csv content differs\nExpected: {expected!r}\nActual:   {actual!r}")


TASK_190 = Task(
    id="task_190_concat_dedupe_sort",
    name="Concat 3 CSVs, dedupe, sort by id",
    tags=("csv", "compute", "pipeline", "execute", "hard"),
    prompt=(
        "В каталоге data три CSV-файла part_a.csv, part_b.csv, part_c.csv с"
        " одинаковыми колонками 'id,name'. Объедини их в один файл merged.csv,"
        " удалив одинаковые строки (где совпадает и id, и name), и отсортируй"
        " результат по столбцу id по возрастанию. Заголовок 'id,name' должен"
        " быть один."
    ),
    setup_files=_CSVS_190,
    gold_files={"merged.csv": _CONCAT_GOLD},
    verifier=_verify_task_190,
)


# 191. sqlite_complex_query_to_csv
def _sqlite_191_setup(ws: Path) -> None:
    conn = sqlite3.connect(ws / "shop.db")
    conn.executescript(
        """
        CREATE TABLE products (id INTEGER PRIMARY KEY, name TEXT, price INTEGER);
        CREATE TABLE orders (id INTEGER PRIMARY KEY, product_id INTEGER, qty INTEGER);
        INSERT INTO products VALUES (1,'A',100),(2,'B',50),(3,'C',200),(4,'D',30);
        INSERT INTO orders (product_id, qty) VALUES (1,3),(2,5),(1,2),(3,1),(4,10),(2,4);
        """
    )
    conn.commit()
    conn.close()


# revenue per product (sum of qty*price): A: (3+2)*100=500, B: (5+4)*50=450, C: 1*200=200, D: 10*30=300
# Sort by revenue desc:
_REV_GOLD_LINES = ["name,revenue", "A,500", "B,450", "D,300", "C,200"]
_REV_GOLD = "\n".join(_REV_GOLD_LINES) + "\n"
TASK_191 = Task(
    id="task_191_sqlite_revenue_report",
    name="Compute revenue per product from shop.db, sort desc, save CSV",
    tags=("sqlite", "csv", "pipeline", "execute", "hard"),
    prompt=(
        "В файле shop.db лежат таблицы:\n"
        "  - products(id, name, price)\n"
        "  - orders(id, product_id, qty)\n"
        "Посчитай для каждого продукта суммарную выручку — sum(qty * price)"
        " по всем заказам. Сохрани результат в revenue.csv с заголовком"
        " 'name,revenue', строки — по убыванию revenue. При равенстве выручки"
        " сортируй по name по возрастанию (в данных такого не случится, но на"
        " всякий случай)."
    ),
    setup_files={},
    setup_callback=_sqlite_191_setup,
    gold_files={"revenue.csv": _REV_GOLD},
    verifier=file_text_equals("revenue.csv", _REV_GOLD),
)


# 192. project_replace_imports
_IMPORTS_FILES_192 = {
    "code/a.py": "from old_pkg.module import foo\n\n\nprint(foo())\n",
    "code/b.py": "from old_pkg.module import bar\nimport old_pkg.helpers as h\n\n\nprint(bar(), h)\n",
    "code/c.py": "import other\n\n\nprint(other)\n",
}
_IMPORTS_RESULT_192 = {
    "code/a.py": "from new_pkg.module import foo\n\n\nprint(foo())\n",
    "code/b.py": "from new_pkg.module import bar\nimport new_pkg.helpers as h\n\n\nprint(bar(), h)\n",
    "code/c.py": "import other\n\n\nprint(other)\n",
}
TASK_192 = Task(
    id="task_192_rewrite_imports",
    name="Rewrite `old_pkg` → `new_pkg` across code/",
    tags=("refactor", "python", "multifile", "execute", "hard"),
    prompt=(
        "В каталоге code три .py-файла: a.py, b.py, c.py. В файлах a.py и"
        " b.py есть импорты, начинающиеся с 'old_pkg' (либо 'from old_pkg.X"
        " import ...', либо 'import old_pkg.X as ...'). Замени везде 'old_pkg'"
        " на 'new_pkg' в импортах. Файл c.py — без изменений. Остальной код"
        " (после импортов) не трогай."
    ),
    setup_files=_IMPORTS_FILES_192,
    gold_files=_IMPORTS_RESULT_192,
    verifier=all_of(
        file_not_contains("code/a.py", "old_pkg"),
        file_not_contains("code/b.py", "old_pkg"),
        file_contains("code/a.py", "new_pkg.module"),
        file_contains("code/b.py", "new_pkg.module", "new_pkg.helpers"),
        file_text_equals("code/c.py", _IMPORTS_FILES_192["code/c.py"]),
    ),
)


# 193. impl_decorator_memoize
TASK_193 = Task(
    id="task_193_impl_memoize",
    name="Implement a memoize decorator so pytest passes",
    tags=("python", "impl", "pytest", "execute", "hard"),
    prompt=(
        "Создай файл memoize.py с декоратором memoize(func), который"
        " кэширует возвращаемые значения по позиционным аргументам. Повторный"
        " вызов с тем же набором аргументов должен вернуть сохранённое"
        " значение без повторного вызова обёрнутой функции. Тесты в"
        " tests/test_memoize.py должны пройти."
    ),
    setup_files={
        "tests/test_memoize.py": (
            "from memoize import memoize\n"
            "\n"
            "\n"
            "def _build():\n"
            "    state = {'n': 0}\n"
            "\n"
            "    @memoize\n"
            "    def add(a, b):\n"
            "        state['n'] += 1\n"
            "        return a + b\n"
            "\n"
            "    return add, state\n"
            "\n"
            "\n"
            "def test_basic_returns():\n"
            "    add, _ = _build()\n"
            "    assert add(1, 2) == 3\n"
            "    assert add(3, 4) == 7\n"
            "\n"
            "\n"
            "def test_caches_result():\n"
            "    add, state = _build()\n"
            "    add(10, 20)\n"
            "    add(10, 20)\n"
            "    add(10, 20)\n"
            "    assert state['n'] == 1\n"
            "\n"
            "\n"
            "def test_different_args_recompute():\n"
            "    add, state = _build()\n"
            "    add(1, 1)\n"
            "    add(1, 2)\n"
            "    add(2, 1)\n"
            "    assert state['n'] == 3\n"
        ),
    },
    gold_files={
        "memoize.py": (
            "def memoize(func):\n"
            "    cache = {}\n"
            "\n"
            "    def wrapper(*args):\n"
            "        if args not in cache:\n"
            "            cache[args] = func(*args)\n"
            "        return cache[args]\n"
            "\n"
            "    return wrapper\n"
        ),
    },
    verifier=pytest_passes("tests"),
)


# 194. impl_context_manager
TASK_194 = Task(
    id="task_194_impl_context_manager",
    name="Implement Timer context manager so pytest passes",
    tags=("python", "impl", "pytest", "execute", "hard"),
    prompt=(
        "Создай файл timer.py с классом Timer, который работает как контекстный"
        " менеджер: при входе в блок with запоминает текущее время, при"
        " выходе — записывает прошедшее время (число секунд, любого типа —"
        " float или int) в атрибут elapsed. Тесты в tests/test_timer.py"
        " должны пройти."
    ),
    setup_files={
        "tests/test_timer.py": (
            "import time\n"
            "\n"
            "from timer import Timer\n"
            "\n"
            "\n"
            "def test_records_elapsed():\n"
            "    with Timer() as t:\n"
            "        time.sleep(0.01)\n"
            "    assert t.elapsed >= 0\n"
            "    assert t.elapsed < 1\n"
            "\n"
            "\n"
            "def test_can_be_reused():\n"
            "    t = Timer()\n"
            "    with t:\n"
            "        pass\n"
            "    first = t.elapsed\n"
            "    with t:\n"
            "        time.sleep(0.01)\n"
            "    assert t.elapsed >= first\n"
        ),
    },
    gold_files={
        "timer.py": (
            "import time\n"
            "\n"
            "\n"
            "class Timer:\n"
            "    def __enter__(self):\n"
            "        self._start = time.monotonic()\n"
            "        return self\n"
            "\n"
            "    def __exit__(self, exc_type, exc, tb):\n"
            "        self.elapsed = time.monotonic() - self._start\n"
            "        return False\n"
        ),
    },
    verifier=pytest_passes("tests"),
)


# 195. impl_iterator
TASK_195 = Task(
    id="task_195_impl_iterator",
    name="Implement Range-like iterator class so pytest passes",
    tags=("python", "impl", "pytest", "execute", "medium"),
    prompt=(
        "Создай файл range_like.py с классом MyRange(start, stop, step=1)."
        " Он должен поддерживать итерацию: list(MyRange(0, 5)) == [0, 1, 2, 3,"
        " 4]; list(MyRange(10, 0, -2)) == [10, 8, 6, 4, 2]. Тесты в"
        " tests/test_range.py должны пройти."
    ),
    setup_files={
        "tests/test_range.py": (
            "from range_like import MyRange\n"
            "\n"
            "\n"
            "def test_forward():\n"
            "    assert list(MyRange(0, 5)) == [0, 1, 2, 3, 4]\n"
            "\n"
            "\n"
            "def test_backward():\n"
            "    assert list(MyRange(10, 0, -2)) == [10, 8, 6, 4, 2]\n"
            "\n"
            "\n"
            "def test_empty():\n"
            "    assert list(MyRange(5, 5)) == []\n"
            "\n"
            "\n"
            "def test_step_two():\n"
            "    assert list(MyRange(0, 6, 2)) == [0, 2, 4]\n"
        ),
    },
    gold_files={
        "range_like.py": (
            "class MyRange:\n"
            "    def __init__(self, start, stop, step=1):\n"
            "        self.start = start\n"
            "        self.stop = stop\n"
            "        self.step = step\n"
            "\n"
            "    def __iter__(self):\n"
            "        v = self.start\n"
            "        if self.step > 0:\n"
            "            while v < self.stop:\n"
            "                yield v\n"
            "                v += self.step\n"
            "        elif self.step < 0:\n"
            "            while v > self.stop:\n"
            "                yield v\n"
            "                v += self.step\n"
        ),
    },
    verifier=pytest_passes("tests"),
)


# ---------------------------------------------------------------------------
# Group H: Final hard (5 tasks, 196..200)
# ---------------------------------------------------------------------------

# 196. xlsx_to_multi_format
def _xlsx_196_setup(ws: Path) -> None:
    import openpyxl  # noqa: PLC0415

    wb = openpyxl.Workbook()
    sh = wb.active
    sh.title = "Data"
    sh.append(["name", "value"])
    for n, v in [("alpha", 1), ("bravo", 2), ("charlie", 3)]:
        sh.append([n, v])
    wb.save(ws / "data.xlsx")


_MULTI_CSV = "name,value\nalpha,1\nbravo,2\ncharlie,3\n"
_MULTI_JSON_DATA = [
    {"name": "alpha", "value": 1},
    {"name": "bravo", "value": 2},
    {"name": "charlie", "value": 3},
]


def _verify_task_196(ws: Path) -> VerifyResult:
    csv_p = ws / "data.csv"
    json_p = ws / "data.json"
    if not csv_p.exists():
        return VerifyResult(False, "data.csv missing")
    if not json_p.exists():
        return VerifyResult(False, "data.json missing")
    if csv_p.read_text().strip() != _MULTI_CSV.strip():
        return VerifyResult(False, f"data.csv content differs from expected\nGot:\n{csv_p.read_text()}")
    try:
        actual = json.loads(json_p.read_text())
    except json.JSONDecodeError as exc:
        return VerifyResult(False, f"data.json invalid: {exc}")
    expected_by_name = {r["name"]: r["value"] for r in _MULTI_JSON_DATA}
    if not isinstance(actual, list):
        return VerifyResult(False, f"data.json must be a list, got {type(actual).__name__}")
    actual_by_name = {row.get("name"): row.get("value") for row in actual if isinstance(row, dict)}
    if actual_by_name != expected_by_name:
        return VerifyResult(False, f"data.json {actual_by_name!r} differs from {expected_by_name!r}")
    return VerifyResult(True, "data.csv and data.json both reflect the xlsx contents")


TASK_196 = Task(
    id="task_196_xlsx_to_csv_and_json",
    name="Export data.xlsx to both data.csv and data.json",
    tags=("xlsx", "csv", "json", "convert", "execute", "hard"),
    prompt=(
        "В файле data.xlsx (лист Data) колонки 'name' и 'value', три строки"
        " данных. Сделай две вещи:\n"
        "  - data.csv: тот же контент в CSV (заголовок 'name,value' + три"
        " строки данных в исходном порядке);\n"
        "  - data.json: JSON-массив объектов с полями name (строка) и value"
        " (число), по одному объекту на строку.\n"
        "Используй openpyxl для чтения xlsx."
    ),
    setup_files={},
    setup_callback=_xlsx_196_setup,
    gold_files={
        "data.csv": _MULTI_CSV,
        "data.json": json.dumps(_MULTI_JSON_DATA) + "\n",
    },
    verifier=_verify_task_196,
)


# 197. md_filter_by_frontmatter
_POSTS = {
    "posts/a.md": "---\ntitle: A\ntags: [python]\n---\nbody\n",
    "posts/b.md": "---\ntitle: B\ntags: [js]\n---\nbody\n",
    "posts/c.md": "---\ntitle: C\ntags: [python, web]\n---\nbody\n",
    "posts/d.md": "---\ntitle: D\ntags: [rust]\n---\nbody\n",
    "posts/e.md": "---\ntitle: E\ntags: [python, db]\n---\nbody\n",
}
_PY_POSTS = sorted({"a.md", "c.md", "e.md"})


def _verify_task_197(ws: Path) -> VerifyResult:
    p = ws / "python_posts.txt"
    if not p.exists():
        return VerifyResult(False, "python_posts.txt missing")
    raw = sorted({line.strip().split("/")[-1] for line in p.read_text().splitlines() if line.strip()})
    if raw == _PY_POSTS:
        return VerifyResult(True, "python_posts.txt has the three python-tagged posts")
    return VerifyResult(False, f"python_posts.txt {raw} differs from {_PY_POSTS}")


TASK_197 = Task(
    id="task_197_filter_md_by_tag",
    name="List posts in posts/ tagged with `python`",
    tags=("markdown", "yaml", "search", "execute", "hard"),
    prompt=(
        "В каталоге posts пять .md-файлов. У каждого в начале — YAML-фронт"
        " матер между двумя строками '---' с полями title и tags (tags — это"
        " список строк). Сделай файл python_posts.txt со списком имён файлов"
        " (без префикса 'posts/' или с ним — оба варианта допустимы), у"
        " которых в tags есть значение 'python'. По одному имени на строку,"
        " порядок не важен."
    ),
    setup_files=_POSTS,
    gold_files={"python_posts.txt": "\n".join(_PY_POSTS) + "\n"},
    verifier=_verify_task_197,
)


# 198. tar_extract — extract a tar archive
def _tar_198_setup(ws: Path) -> None:
    bundle = ws / "bundle.tar.gz"
    with tarfile.open(bundle, "w:gz") as tar:
        for name, content in [
            ("first.txt", "first\n"),
            ("second.txt", "second\n"),
            ("third.txt", "third\n"),
        ]:
            data = content.encode()
            info = tarfile.TarInfo(name=name)
            info.size = len(data)
            tar.addfile(info, io.BytesIO(data))


def _verify_task_198(ws: Path) -> VerifyResult:
    extracted = ws / "extracted"
    if not extracted.is_dir():
        return VerifyResult(False, "extracted/ directory missing")
    for name, content in [("first.txt", "first"), ("second.txt", "second"), ("third.txt", "third")]:
        p = extracted / name
        if not p.exists():
            return VerifyResult(False, f"extracted/{name} missing")
        if p.read_text().strip() != content:
            return VerifyResult(False, f"extracted/{name} content differs")
    return VerifyResult(True, "bundle.tar.gz extracted to extracted/ with all three files")


def _tar_198_gold_callback(ws: Path) -> None:
    extracted = ws / "extracted"
    extracted.mkdir(exist_ok=True)
    (extracted / "first.txt").write_text("first\n")
    (extracted / "second.txt").write_text("second\n")
    (extracted / "third.txt").write_text("third\n")


TASK_198 = Task(
    id="task_198_tar_extract",
    name="Extract bundle.tar.gz into extracted/",
    tags=("archive", "tar", "execute", "hard"),
    prompt=(
        "В корне рабочей директории есть архив bundle.tar.gz (tar+gzip)."
        " Распакуй его так, чтобы файлы first.txt, second.txt, third.txt"
        " оказались в каталоге extracted/. Содержимое файлов сохрани"
        " как есть."
    ),
    setup_files={},
    setup_callback=_tar_198_setup,
    gold_callback=_tar_198_gold_callback,
    verifier=_verify_task_198,
)


# 199. sqlite_to_csv_filtered_query
def _sqlite_199_setup(ws: Path) -> None:
    conn = sqlite3.connect(ws / "events.db")
    conn.executescript(
        """
        CREATE TABLE events (id INTEGER PRIMARY KEY, kind TEXT, value INTEGER);
        INSERT INTO events (kind, value) VALUES
            ('click', 10), ('view', 5), ('click', 20), ('buy', 100),
            ('click', 50), ('view', 1), ('buy', 200), ('click', 30),
            ('view', 0), ('buy', 50);
        """
    )
    conn.commit()
    conn.close()


# clicks with value > 15: (20), (50), (30). Sorted ascending = 20, 30, 50.
_BIG_CLICK_GOLD = "value\n20\n30\n50\n"
TASK_199 = Task(
    id="task_199_sqlite_filtered_export",
    name="Export click events with value>15 from sqlite to CSV",
    tags=("sqlite", "csv", "pipeline", "execute", "hard"),
    prompt=(
        "В файле events.db таблица events(id, kind, value). Выбери из неё"
        " только события с kind == 'click' и value > 15 и сохрани их value в"
        " файл big_clicks.csv с заголовком 'value' — по одной строке на"
        " значение, отсортированные по возрастанию."
    ),
    setup_files={},
    setup_callback=_sqlite_199_setup,
    gold_files={"big_clicks.csv": _BIG_CLICK_GOLD},
    verifier=file_text_equals("big_clicks.csv", _BIG_CLICK_GOLD),
)


# 200. impl_priority_queue
TASK_200 = Task(
    id="task_200_impl_priority_queue",
    name="Implement PriorityQueue class so pytest passes",
    tags=("python", "impl", "pytest", "execute", "hard"),
    prompt=(
        "Создай файл pq.py с классом PriorityQueue. Методы:\n"
        "  - push(item, priority): добавить элемент с приоритетом (любое"
        " число; меньшее число — выше приоритет);\n"
        "  - pop(): извлечь и вернуть элемент с наивысшим приоритетом (то"
        " есть с наименьшим числовым значением приоритета); поднимать"
        " IndexError если пусто;\n"
        "  - __len__(): количество элементов в очереди.\n"
        " Если приоритеты равны, порядок извлечения может быть любым. Тесты в"
        " tests/test_pq.py должны пройти."
    ),
    setup_files={
        "tests/test_pq.py": (
            "import pytest\n"
            "\n"
            "from pq import PriorityQueue\n"
            "\n"
            "\n"
            "def test_basic_order():\n"
            "    q = PriorityQueue()\n"
            "    q.push('b', 2)\n"
            "    q.push('a', 1)\n"
            "    q.push('c', 3)\n"
            "    assert q.pop() == 'a'\n"
            "    assert q.pop() == 'b'\n"
            "    assert q.pop() == 'c'\n"
            "\n"
            "\n"
            "def test_len():\n"
            "    q = PriorityQueue()\n"
            "    assert len(q) == 0\n"
            "    q.push('x', 1)\n"
            "    q.push('y', 0)\n"
            "    assert len(q) == 2\n"
            "    q.pop()\n"
            "    assert len(q) == 1\n"
            "\n"
            "\n"
            "def test_empty_raises():\n"
            "    q = PriorityQueue()\n"
            "    with pytest.raises(IndexError):\n"
            "        q.pop()\n"
        ),
    },
    gold_files={
        "pq.py": (
            "import heapq\n"
            "\n"
            "\n"
            "class PriorityQueue:\n"
            "    def __init__(self):\n"
            "        self._heap = []\n"
            "        self._counter = 0\n"
            "\n"
            "    def push(self, item, priority):\n"
            "        heapq.heappush(self._heap, (priority, self._counter, item))\n"
            "        self._counter += 1\n"
            "\n"
            "    def pop(self):\n"
            "        if not self._heap:\n"
            "            raise IndexError('pop from empty PriorityQueue')\n"
            "        return heapq.heappop(self._heap)[2]\n"
            "\n"
            "    def __len__(self):\n"
            "        return len(self._heap)\n"
        ),
    },
    verifier=pytest_passes("tests"),
)


# ---------------------------------------------------------------------------
# Group I: Extra-hard extensions (5 tasks, 201..205)
# ---------------------------------------------------------------------------

# 201. multi-source reconciliation (CSV + JSONL + blacklist)
_USERS_201 = (
    "user_id,email\n"
    "1,alice@example.com\n"
    "2,bob@example.com\n"
    "3,carol@example.com\n"
    "4,dave@example.com\n"
    "5,erin@example.com\n"
    "6,fraud@example.com\n"
)
_ORDERS_201 = [
    {"user_id": 1, "amount": 120, "status": "paid"},
    {"user_id": 1, "amount": 200, "status": "paid"},
    {"user_id": 1, "amount": 70, "status": "refund"},
    {"user_id": 2, "amount": 150, "status": "paid"},
    {"user_id": 2, "amount": 170, "status": "paid"},
    {"user_id": 3, "amount": 500, "status": "paid"},
    {"user_id": 4, "amount": 100, "status": "paid"},
    {"user_id": 4, "amount": 110, "status": "paid"},
    {"user_id": 4, "amount": 120, "status": "paid"},
    {"user_id": 5, "amount": 50, "status": "paid"},
    {"user_id": 5, "amount": 40, "status": "paid"},
    {"user_id": 5, "amount": 30, "status": "paid"},
    {"user_id": 6, "amount": 1000, "status": "paid"},
]
_ORDERS_201_JSONL = "".join(json.dumps(row) + "\n" for row in _ORDERS_201)
_BLACKLIST_201 = "carol@example.com\nfraud@example.com\n"
_VIP_201_GOLD = [
    {"user_id": 4, "email": "dave@example.com", "paid_total": 330, "paid_orders": 3},
    {"user_id": 1, "email": "alice@example.com", "paid_total": 320, "paid_orders": 2},
    {"user_id": 2, "email": "bob@example.com", "paid_total": 320, "paid_orders": 2},
]
TASK_201 = Task(
    id="task_201_reconcile_vip_users",
    name="Reconcile users/orders/blacklist into vip_users.json",
    tags=("csv", "jsonl", "pipeline", "execute", "hard"),
    prompt=(
        "В рабочей директории есть три файла:\n"
        "  - users.csv (user_id,email),\n"
        "  - orders.jsonl (user_id,amount,status),\n"
        "  - blacklist.txt (по одному email на строку).\n"
        "Собери vip_users.json — JSON-массив объектов с полями"
        " user_id, email, paid_total, paid_orders для пользователей, которые:\n"
        "  1) НЕ находятся в blacklist.txt,\n"
        "  2) имеют paid_total >= 300 (сумма amount только по статусу paid),\n"
        "  3) имеют paid_orders >= 2 (количество строк со статусом paid).\n"
        "Отсортируй массив по paid_total по убыванию, при равенстве — по email"
        " по возрастанию."
    ),
    setup_files={
        "users.csv": _USERS_201,
        "orders.jsonl": _ORDERS_201_JSONL,
        "blacklist.txt": _BLACKLIST_201,
    },
    gold_files={"vip_users.json": json.dumps(_VIP_201_GOLD, ensure_ascii=False) + "\n"},
    verifier=_json_file_matches_loose("vip_users.json", _VIP_201_GOLD, ordered=True),
)


# 202. nested zip extraction + schema normalization + aggregate
def _zip_202_setup(ws: Path) -> None:
    with zipfile.ZipFile(ws / "datasets.zip", "w") as zf:
        zf.writestr("january/sales.csv", "region,amount\neu,100\nus,200\napac,150\n")
        zf.writestr("february/sales.csv", "amount,region\n120,eu\n220,us\n130,apac\n")
        zf.writestr("march/sales.csv", "region,amount\neu,90\nus,210\nlatam,80\n")


_REGION_202_GOLD = (
    "region,total\n"
    "us,630\n"
    "eu,310\n"
    "apac,280\n"
    "latam,80\n"
)
TASK_202 = Task(
    id="task_202_zip_sales_consolidation",
    name="Extract datasets.zip and consolidate monthly sales",
    tags=("archive", "csv", "pipeline", "execute", "hard"),
    prompt=(
        "В файле datasets.zip лежат три CSV-файла (по одному в папках"
        " january/february/march). В двух файлах колонки order: region,amount,"
        " а в одном amount,region. Твоя задача:\n"
        "  1) распаковать архив,\n"
        "  2) объединить данные,\n"
        "  3) посчитать сумму amount по region,\n"
        "  4) сохранить результат в region_totals.csv (заголовок region,total),"
        "     отсортировав по total по убыванию,\n"
        "  5) сохранить top_region.txt с названием региона с максимальной суммой."
    ),
    setup_files={},
    setup_callback=_zip_202_setup,
    gold_files={"region_totals.csv": _REGION_202_GOLD, "top_region.txt": "us\n"},
    verifier=all_of(
        file_text_equals("region_totals.csv", _REGION_202_GOLD),
        file_text_equals("top_region.txt", "us"),
    ),
)


# 203. sqlite -> markdown analytical report
def _sqlite_203_setup(ws: Path) -> None:
    conn = sqlite3.connect(ws / "support.db")
    conn.executescript(
        """
        CREATE TABLE tickets (id INTEGER PRIMARY KEY, team TEXT, status TEXT, hours INTEGER);
        INSERT INTO tickets (team, status, hours) VALUES
            ('alpha', 'closed', 3),
            ('alpha', 'open', 2),
            ('alpha', 'closed', 5),
            ('beta', 'open', 4),
            ('beta', 'open', 1),
            ('beta', 'closed', 2),
            ('gamma', 'closed', 7),
            ('gamma', 'closed', 1);
        """
    )
    conn.commit()
    conn.close()


_REPORT_203_GOLD = (
    "| team | open | closed | closed_hours |\n"
    "| --- | --- | --- | --- |\n"
    "| alpha | 1 | 2 | 8 |\n"
    "| beta | 2 | 1 | 2 |\n"
    "| gamma | 0 | 2 | 8 |\n"
    "\n"
    "TOTAL_OPEN=3\n"
)
TASK_203 = Task(
    id="task_203_sqlite_team_markdown_report",
    name="Build markdown KPI report from support.db tickets",
    tags=("sqlite", "markdown", "pipeline", "execute", "hard"),
    prompt=(
        "В support.db есть таблица tickets(id, team, status, hours)."
        " Построй отчет team_report.md в виде markdown-таблицы с колонками"
        " team | open | closed | closed_hours, где:\n"
        "  - open = количество тикетов status='open',\n"
        "  - closed = количество тикетов status='closed',\n"
        "  - closed_hours = сумма hours только по status='closed'.\n"
        "Строки отсортируй по team по возрастанию. После таблицы добавь пустую"
        " строку и строку вида TOTAL_OPEN=<число>."
    ),
    setup_files={},
    setup_callback=_sqlite_203_setup,
    gold_files={"team_report.md": _REPORT_203_GOLD},
    verifier=file_text_equals("team_report.md", _REPORT_203_GOLD),
)


# 204. multi-file API migration + report
_MIGRATION_204_FILES = {
    "service/a.py": (
        "import requests\n"
        "\n"
        "\n"
        "def fetch_user(uid):\n"
        "    return requests.get(f'https://api.example.com/users/{uid}').json()\n"
    ),
    "service/b.py": (
        "import requests as rq\n"
        "\n"
        "\n"
        "def create(payload):\n"
        "    return rq.post('https://api.example.com/create', json=payload)\n"
    ),
    "service/c.py": (
        "from requests import get\n"
        "\n"
        "\n"
        "def ping():\n"
        "    return get('https://api.example.com/ping').status_code\n"
    ),
    "infra/http_client.py": (
        "class HTTPClient:\n"
        "    def get(self, url, **kwargs):\n"
        "        raise NotImplementedError\n"
        "\n"
        "    def post(self, url, **kwargs):\n"
        "        raise NotImplementedError\n"
        "\n"
        "\n"
        "http_client = HTTPClient()\n"
    ),
}


def _verify_task_204(ws: Path) -> VerifyResult:
    checks = [
        ("service/a.py", ["from infra.http_client import http_client", "http_client.get("], ["requests.get", "import requests"]),
        ("service/b.py", ["from infra.http_client import http_client", "http_client.post("], ["rq.post", "import requests as rq"]),
        ("service/c.py", ["from infra.http_client import http_client", "http_client.get("], ["from requests import get", " get("]),
    ]
    for rel, required, forbidden in checks:
        p = ws / rel
        if not p.exists():
            return VerifyResult(False, f"{rel} missing")
        text = p.read_text()
        missing = [s for s in required if s not in text]
        if missing:
            return VerifyResult(False, f"{rel} missing required snippets: {missing}")
        present = [s for s in forbidden if s in text]
        if present:
            return VerifyResult(False, f"{rel} still has legacy API snippets: {present}")
    report = ws / "migration_report.json"
    if not report.exists():
        return VerifyResult(False, "migration_report.json missing")
    try:
        data = json.loads(report.read_text())
    except json.JSONDecodeError as exc:
        return VerifyResult(False, f"migration_report.json invalid JSON: {exc}")
    if data.get("files_updated") != 3 or data.get("replacements") != 3:
        return VerifyResult(False, f"migration_report.json mismatch: {data!r}")
    return VerifyResult(True, "requests API migrated in 3 files and report is correct")


TASK_204 = Task(
    id="task_204_requests_to_http_client_migration",
    name="Migrate 3 files from requests calls to infra.http_client",
    tags=("refactor", "python", "multifile", "execute", "hard"),
    prompt=(
        "В файлах service/a.py, service/b.py, service/c.py есть вызовы"
        " requests API (в разных формах импорта). Выполни миграцию на"
        " централизованный клиент из infra/http_client.py:\n"
        "  - в каждом из трех service/*.py должен быть импорт"
        "    'from infra.http_client import http_client';\n"
        "  - requests.get(...) заменить на http_client.get(...);\n"
        "  - requests.post(...)/rq.post(...) заменить на http_client.post(...);\n"
        "  - from requests import get + get(...) тоже мигрировать на"
        "    http_client.get(...).\n"
        "URL и аргументы вызовов не менять. После миграции создай"
        " migration_report.json с объектом {'files_updated': 3, 'replacements': 3}."
    ),
    setup_files=_MIGRATION_204_FILES,
    gold_files={
        "service/a.py": (
            "from infra.http_client import http_client\n"
            "\n"
            "\n"
            "def fetch_user(uid):\n"
            "    return http_client.get(f'https://api.example.com/users/{uid}').json()\n"
        ),
        "service/b.py": (
            "from infra.http_client import http_client\n"
            "\n"
            "\n"
            "def create(payload):\n"
            "    return http_client.post('https://api.example.com/create', json=payload)\n"
        ),
        "service/c.py": (
            "from infra.http_client import http_client\n"
            "\n"
            "\n"
            "def ping():\n"
            "    return http_client.get('https://api.example.com/ping').status_code\n"
        ),
        "migration_report.json": '{"files_updated": 3, "replacements": 3}\n',
    },
    verifier=_verify_task_204,
)


# 205. advanced pytest implementation task
TASK_205 = Task(
    id="task_205_impl_paid_aggregator",
    name="Implement normalize_transactions so pytest passes",
    tags=("python", "impl", "pytest", "execute", "hard"),
    prompt=(
        "Создай файл pipeline.py с функцией normalize_transactions(rows)."
        " Вход: список словарей с полями user, amount, status.\n"
        " Поведение:\n"
        "  - учитывай только строки со status == 'paid';\n"
        "  - amount может быть int/float/str, нужно корректно привести к float;\n"
        "  - если amount нельзя привести к числу у paid-строки — бросай ValueError;\n"
        "  - агрегируй сумму amount по user;\n"
        "  - верни list[tuple[user, total]], где total округлен до 2 знаков;\n"
        "  - сортировка: total по убыванию, при равенстве user по возрастанию.\n"
        "Тесты в tests/test_pipeline.py должны пройти."
    ),
    setup_files={
        "tests/test_pipeline.py": (
            "import pytest\n"
            "\n"
            "from pipeline import normalize_transactions\n"
            "\n"
            "\n"
            "def test_basic_aggregation_and_sort():\n"
            "    rows = [\n"
            "        {'user': 'alice', 'amount': '100.20', 'status': 'paid'},\n"
            "        {'user': 'bob', 'amount': 50, 'status': 'paid'},\n"
            "        {'user': 'alice', 'amount': 49.8, 'status': 'paid'},\n"
            "        {'user': 'bob', 'amount': 70, 'status': 'refund'},\n"
            "        {'user': 'carol', 'amount': '75', 'status': 'paid'},\n"
            "    ]\n"
            "    assert normalize_transactions(rows) == [('alice', 150.0), ('carol', 75.0), ('bob', 50.0)]\n"
            "\n"
            "\n"
            "def test_tie_breaker_by_user_name():\n"
            "    rows = [\n"
            "        {'user': 'zoe', 'amount': 10, 'status': 'paid'},\n"
            "        {'user': 'amy', 'amount': 10, 'status': 'paid'},\n"
            "    ]\n"
            "    assert normalize_transactions(rows) == [('amy', 10.0), ('zoe', 10.0)]\n"
            "\n"
            "\n"
            "def test_ignores_non_paid():\n"
            "    rows = [\n"
            "        {'user': 'alice', 'amount': 100, 'status': 'refund'},\n"
            "        {'user': 'bob', 'amount': 50, 'status': 'failed'},\n"
            "    ]\n"
            "    assert normalize_transactions(rows) == []\n"
            "\n"
            "\n"
            "def test_invalid_paid_amount_raises():\n"
            "    rows = [\n"
            "        {'user': 'alice', 'amount': 'oops', 'status': 'paid'},\n"
            "    ]\n"
            "    with pytest.raises(ValueError):\n"
            "        normalize_transactions(rows)\n"
            "\n"
            "\n"
            "def test_rounding_to_2_digits():\n"
            "    rows = [\n"
            "        {'user': 'alice', 'amount': '0.3333', 'status': 'paid'},\n"
            "        {'user': 'alice', 'amount': '0.3333', 'status': 'paid'},\n"
            "        {'user': 'alice', 'amount': '0.3333', 'status': 'paid'},\n"
            "    ]\n"
            "    assert normalize_transactions(rows) == [('alice', 1.0)]\n"
        ),
    },
    gold_files={
        "pipeline.py": (
            "def normalize_transactions(rows):\n"
            "    totals = {}\n"
            "    for row in rows:\n"
            "        if row.get('status') != 'paid':\n"
            "            continue\n"
            "        user = row.get('user')\n"
            "        try:\n"
            "            amount = float(row.get('amount'))\n"
            "        except (TypeError, ValueError):\n"
            "            raise ValueError('invalid amount for paid transaction')\n"
            "        totals[user] = totals.get(user, 0.0) + amount\n"
            "    out = [(user, round(total, 2)) for user, total in totals.items()]\n"
            "    out.sort(key=lambda x: (-x[1], x[0]))\n"
            "    return out\n"
        ),
    },
    verifier=pytest_passes("tests"),
)


EXTREME_TASKS: list[Task] = [
    TASK_151, TASK_152, TASK_153, TASK_154, TASK_155, TASK_156, TASK_157, TASK_158,
    TASK_159, TASK_160, TASK_161,
    TASK_162, TASK_163, TASK_164, TASK_165, TASK_166,
    TASK_167, TASK_168, TASK_169, TASK_170, TASK_171, TASK_172, TASK_173, TASK_174,
    TASK_175, TASK_176, TASK_177, TASK_178, TASK_179, TASK_180, TASK_181, TASK_182,
    TASK_183, TASK_184, TASK_185, TASK_186, TASK_187,
    TASK_188, TASK_189, TASK_190, TASK_191, TASK_192, TASK_193, TASK_194, TASK_195,
    TASK_196, TASK_197, TASK_198, TASK_199, TASK_200,
    TASK_201, TASK_202, TASK_203, TASK_204, TASK_205,
]


# Keep yaml/sqlite_query_returns/xlsx_cell_equals/file_exists/file_lines_equal
# imported even if not every task uses them directly — they're useful
# building blocks for variants and ruff would otherwise prune the imports.
_ = (yaml, sqlite_query_returns, xlsx_cell_equals, file_exists, file_lines_equal)
